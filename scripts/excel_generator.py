"""
Excel Generator Module for Warehouse Analysis Tool V2

PURPOSE:
This module generates comprehensive Excel reports from analysis results.
It creates professional, multi-sheet workbooks with:
- Executive summary and key metrics
- Detailed analysis results from all modules
- Charts and visualizations
- Raw data preservation
- Configuration documentation

FOR BEGINNERS:
- This module takes analysis results and creates Excel files
- Each analysis gets its own sheet with tables and summaries
- Professional formatting makes reports presentation-ready
- Charts and graphs help visualize key insights
- All configuration settings are documented for reproducibility
"""

import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import warnings
import sys
from pathlib import Path
import io
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Import configuration
sys.path.append(str(Path(__file__).parent.parent))
import config

class ExcelGenerator:
    """
    Excel report generation class.
    
    This class creates comprehensive Excel reports including:
    - Executive summary with key metrics
    - Detailed analysis results from all modules
    - Professional formatting and styling
    - Charts and visualizations
    - Configuration documentation
    """
    
    def __init__(self, analysis_results, configuration=None, output_settings=None):
        """
        Initialize the ExcelGenerator.
        
        Args:
            analysis_results (dict): Combined results from all analysis modules
            configuration (dict): Analysis configuration settings
            output_settings (dict): Output formatting preferences
        """
        self.analysis_results = analysis_results
        self.configuration = configuration or {}
        self.output_settings = output_settings or {}
        
        # Set formatting preferences
        self.currency_symbol = self.output_settings.get('CURRENCY_SYMBOL', '$')
        self.decimal_places = self.output_settings.get('DECIMAL_PLACES', 2)
        self.verbose = self.output_settings.get('VERBOSE_OUTPUT', False)
        
        # Initialize Excel buffer
        self.excel_buffer = None
        
        if self.verbose:
            print("ExcelGenerator initialized")
            print(f"Available analysis results: {list(self.analysis_results.keys())}")
    
    def generate_comprehensive_report(self):
        """
        Generate comprehensive Excel report with all analysis results.
        
        Returns:
            io.BytesIO: Excel file buffer ready for download
        """
        print("📋 Generating comprehensive Excel report...")
        
        # Create Excel buffer
        self.excel_buffer = io.BytesIO()
        
        try:
            with pd.ExcelWriter(self.excel_buffer, engine='openpyxl') as writer:
                
                # Generate each sheet
                self._create_executive_summary(writer)
                self._create_order_analysis_sheet(writer)
                self._create_sku_analysis_sheet(writer)
                self._create_abc_fms_sheet(writer)
                self._create_inventory_analysis_sheet(writer)
                self._create_receipt_analysis_sheet(writer)
                self._create_manpower_analysis_sheet(writer)
                self._create_recommendations_sheet(writer)
                self._create_configuration_sheet(writer)
                self._create_raw_data_summary(writer)
                self._format_workbook(writer)

                if self.verbose:
                    print("All sheets generated successfully")
            
            # Reset buffer position
            self.excel_buffer.seek(0)
            
            print("✅ Excel report generation completed")
            return self.excel_buffer.getvalue()
        
        except Exception as e:
            print(f"❌ Error generating Excel report: {str(e)}")
            raise e
    
    def _format_workbook(self, writer):
        """Apply professional formatting to all sheets in the workbook."""
        HEADER_FILL  = PatternFill('solid', fgColor='1F4E79')   # dark navy
        HEADER_FONT  = Font(bold=True, color='FFFFFF', name='Calibri', size=11)
        SECTION_FILL = PatternFill('solid', fgColor='D6E4F0')   # light blue
        SECTION_FONT = Font(bold=True, color='1F4E79', name='Calibri', size=10)

        for sheet_name, ws in writer.sheets.items():

            # 1. Freeze pane at A2
            ws.freeze_panes = 'A2'

            # 2. Bold + navy header row (row 1 of every sheet)
            for cell in ws[1]:
                if cell.value is not None:
                    cell.font = HEADER_FONT
                    cell.fill = HEADER_FILL
                    cell.alignment = Alignment(horizontal='center', vertical='center')

            # 3. Auto-width all columns (capped at 55)
            for col in ws.columns:
                max_len = max((len(str(cell.value or '')) for cell in col), default=8)
                ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 55)

            # 4. Executive Summary: section header rows + number formatting
            if sheet_name == 'Executive_Summary':
                for row in ws.iter_rows(min_row=2):
                    metric_cell = row[0]
                    value_cell  = row[1] if len(row) > 1 else None
                    metric_val  = str(metric_cell.value or '')
                    value_val   = value_cell.value if value_cell else None

                    is_section = (
                        metric_val == metric_val.upper()
                        and metric_val.strip()
                        and not metric_val.startswith(' ')
                        and (value_val is None or value_val == '')
                    )
                    if is_section:
                        metric_cell.font = SECTION_FONT
                        metric_cell.fill = SECTION_FILL
                        if value_cell:
                            value_cell.fill = SECTION_FILL
                    elif value_cell and isinstance(value_val, float):
                        value_cell.number_format = '#,##0.0'
                    elif value_cell and isinstance(value_val, int):
                        value_cell.number_format = '#,##0'

    def _create_executive_summary(self, writer):
        """Create executive summary sheet with KPIs from all analysis modules."""
        if self.verbose:
            print("Creating Executive Summary sheet...")

        summary_data = []

        # ── 1. METADATA ──────────────────────────────────────────────────────────
        summary_data.append(['Analysis Date', datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
        summary_data.append(['Report Generator', 'Warehouse Analysis Tool V2'])

        if 'data_loader' in self.analysis_results:
            dl = self.analysis_results['data_loader'].get('data_summary', {})
            dr = dl.get('date_range', {})
            summary_data.append(['Analysis Period',
                                 f"{dr.get('start', 'N/A')} to {dr.get('end', 'N/A')} "
                                 f"({dr.get('days', 'N/A')} days)"])
            summary_data.append(['Total Order Records', dl.get('total_records', 'N/A')])
            summary_data.append(['Unique SKUs in Orders', dl.get('unique_skus', 'N/A')])

        summary_data.append(['', ''])

        # ── 2. OUTBOUND OPERATIONS ───────────────────────────────────────────────
        if 'order_analysis' in self.analysis_results:
            order_results = self.analysis_results['order_analysis']
            ob      = order_results.get('outbound_summary', {})
            overall = ob.get('overall_stats', {})
            each_s  = ob.get('each_picks', {})
            case_s  = ob.get('case_picks', {})
            mixed_s = ob.get('mixed_picks', {})

            summary_data.append(['OUTBOUND OPERATIONS', ''])

            da = overall.get('daily_average', {})
            summary_data.append(['  Avg Daily Orders',         round(da.get('orders', 0), 0)])
            summary_data.append(['  Avg Daily Order Lines',    round(da.get('lines',  0), 0)])
            summary_data.append(['  Avg Daily Case Equiv Vol', round(da.get('volume', 0), 1)])

            ap = overall.get('absolute_peak', {})
            summary_data.append(['  Absolute Peak Orders', ap.get('orders', 'N/A')])
            summary_data.append(['  Absolute Peak Lines',  ap.get('lines',  'N/A')])
            summary_data.append(['  Absolute Peak CEV',
                                 round(ap.get('volume', 0), 1) if ap.get('volume') else 'N/A'])

            dp = overall.get('design_peak', {})
            summary_data.append(['  Design Peak Orders (95%ile)',
                                 round(dp.get('orders', 0), 0) if dp.get('orders') else 'N/A'])
            summary_data.append(['  Design Peak Lines (95%ile)',
                                 round(dp.get('lines',  0), 0) if dp.get('lines')  else 'N/A'])
            summary_data.append(['  Design Peak CEV (95%ile)',
                                 round(dp.get('volume', 0), 1) if dp.get('volume') else 'N/A'])

            e_da = each_s.get('daily_average',  {})
            c_da = case_s.get('daily_average',  {})
            m_da = mixed_s.get('daily_average', {})
            summary_data.append(['  Each-Pick Avg Daily CEV',  round(e_da.get('volume', 0), 1)])
            summary_data.append(['  Case-Pick Avg Daily CEV',  round(c_da.get('volume', 0), 1)])
            summary_data.append(['  Mixed-Pick Avg Daily CEV', round(m_da.get('volume', 0), 1)])

            summary_data.append(['', ''])

        # ── 3. SKU PORTFOLIO ─────────────────────────────────────────────────────
        if 'sku_analysis' in self.analysis_results:
            sku_results = self.analysis_results['sku_analysis']
            ps = sku_results.get('sku_performance', {}).get('performance_summary', {})

            summary_data.append(['SKU PORTFOLIO', ''])
            summary_data.append(['  Total SKUs Analyzed',           ps.get('total_skus', 'N/A')])
            summary_data.append(['  Daily Movers (Activity > 50%)', ps.get('daily_movers', 'N/A')])
            summary_data.append(['  Occasional Movers (10-50%)',    ps.get('occasional_movers', 'N/A')])
            summary_data.append(['  Slow Movers (Activity <= 10%)', ps.get('slow_movers', 'N/A')])

            dp_dist = ps.get('demand_pattern_distribution', {})
            if dp_dist:
                summary_data.append(['  Demand Pattern - Smooth',       dp_dist.get('Smooth', 0)])
                summary_data.append(['  Demand Pattern - Erratic',      dp_dist.get('Erratic', 0)])
                summary_data.append(['  Demand Pattern - Intermittent', dp_dist.get('Intermittent', 0)])
                summary_data.append(['  Demand Pattern - Lumpy',        dp_dist.get('Lumpy', 0)])

            lc_dist = ps.get('lifecycle_distribution', {})
            if lc_dist:
                summary_data.append(['  Lifecycle - New',          lc_dist.get('New', 0)])
                summary_data.append(['  Lifecycle - Active',       lc_dist.get('Active', 0)])
                summary_data.append(['  Lifecycle - Declining',    lc_dist.get('Declining', 0)])
                summary_data.append(['  Lifecycle - Dead/Dormant', lc_dist.get('Dead/Dormant', 0)])

            summary_data.append(['', ''])

        # ── 4. INVENTORY HEALTH ──────────────────────────────────────────────────
        if 'inventory_analysis' in self.analysis_results:
            inv_results = self.analysis_results['inventory_analysis']
            inv_stats   = inv_results.get('inventory_statistics', {})
            health      = inv_stats.get('inventory_health', {})
            stock_lvls  = inv_stats.get('stock_levels', {})

            summary_data.append(['INVENTORY HEALTH', ''])
            summary_data.append(['  Total Inventory SKUs',    inv_stats.get('total_skus', 'N/A')])
            summary_data.append(['  Avg Daily Stock (Cases)',
                                 round(stock_lvls.get('avg_daily_cases', 0), 0)
                                 if stock_lvls.get('avg_daily_cases') else 'N/A'])

            if health:
                summary_data.append(['  Avg Inventory Turns (Annual)', health.get('avg_inventory_turns', 'N/A')])
                summary_data.append(['  Avg Days of Supply',           health.get('avg_days_of_supply',   'N/A')])
                summary_data.append(['  SKUs - Low Stock',             health.get('skus_low_stock',        'N/A')])
                summary_data.append(['  SKUs - Excess Stock',          health.get('skus_excess_stock',     'N/A')])
                summary_data.append(['  SKUs - OK',                    health.get('skus_ok',               'N/A')])
                summary_data.append(['  SKUs - No Demand',             health.get('skus_no_demand',        'N/A')])

            summary_data.append(['', ''])

        # ── 5. INBOUND / RECEIPT OPERATIONS ──────────────────────────────────────
        if 'receipt_analysis' in self.analysis_results:
            rec_results = self.analysis_results['receipt_analysis']
            dp_rec    = rec_results.get('daily_patterns', {})
            dock_util = rec_results.get('dock_utilization', {})
            eff       = rec_results.get('receiving_efficiency', {}).get('efficiency_stats', {})
            lt        = rec_results.get('lead_times', {})

            summary_data.append(['INBOUND / RECEIPT OPERATIONS', ''])
            summary_data.append(['  Total Receipt Days', dp_rec.get('total_receipt_days', 'N/A')])
            summary_data.append(['  Avg Daily Case Equiv Received',
                                 round(dp_rec.get('avg_daily_case_equivalent_volume', 0), 1)
                                 if dp_rec.get('avg_daily_case_equivalent_volume') else 'N/A'])
            summary_data.append(['  Avg Daily Trucks',
                                 round(dp_rec.get('avg_daily_trucks', 0), 1)
                                 if dp_rec.get('avg_daily_trucks') else 'N/A'])
            summary_data.append(['  Peak Daily CEV Received',
                                 round(dp_rec.get('peak_daily_volume', 0), 1)
                                 if dp_rec.get('peak_daily_volume') else 'N/A'])

            if dock_util:
                summary_data.append(['  Avg Dock Utilization %',
                                     round(dock_util.get('avg_utilization', 0), 1)])
                summary_data.append(['  Over-Capacity Days',
                                     dock_util.get('over_capacity_days', 0)])

            if eff:
                summary_data.append(['  Avg CEV per Truck',
                                     round(eff.get('avg_case_equivalent_per_truck', 0), 1)])

            if 'avg_inter_receipt_days' in lt:
                summary_data.append(['  Avg Inter-Receipt Interval (Days)',
                                     lt.get('avg_inter_receipt_days', 'N/A')])
                summary_data.append(['  P95 Inter-Receipt Interval (Days)',
                                     lt.get('p95_inter_receipt_days', 'N/A')])

            summary_data.append(['', ''])

        # ── 6. MANPOWER REQUIREMENTS ─────────────────────────────────────────────
        if 'manpower_analysis' in self.analysis_results:
            mp_results  = self.analysis_results['manpower_analysis']
            eff_summary = mp_results.get('efficiency_summary', {})

            summary_data.append(['MANPOWER REQUIREMENTS', ''])
            if eff_summary:
                summary_data.append(['  Total Recommended Staff',
                                     eff_summary.get('total_recommended_staff', 'N/A')])
                summary_data.append(['  Pickers Required',
                                     eff_summary.get('recommended_pickers', 'N/A')])
                summary_data.append(['  Receivers Required',
                                     eff_summary.get('recommended_receivers', 'N/A')])
                summary_data.append(['  Loaders Required',
                                     eff_summary.get('recommended_loaders', 'N/A')])
                summary_data.append(['  Est. Daily Labor Hours',
                                     round(eff_summary.get('estimated_daily_labor_hours', 0), 1)
                                     if eff_summary.get('estimated_daily_labor_hours') else 'N/A'])
            else:
                summary_data.append(['  Status', 'Manpower data available - see Manpower_Analysis sheet'])

            summary_data.append(['', ''])

        # ── 7. ABC-FMS CLASSIFICATION ─────────────────────────────────────────────
        if 'abc_fms_analysis' in self.analysis_results:
            abc_results      = self.analysis_results['abc_fms_analysis']
            segment_analysis = abc_results.get('segment_analysis', {})
            critical_segs    = segment_analysis.get('critical_segments', pd.DataFrame())

            summary_data.append(['ABC-FMS CLASSIFICATION', ''])
            summary_data.append(['  Total Segments',   segment_analysis.get('total_segments', 'N/A')])
            summary_data.append(['  Critical Segments',
                                 len(critical_segs) if not critical_segs.empty else 0])
            if not critical_segs.empty and 'Volume_Percent' in critical_segs.columns:
                summary_data.append(['  Critical Segment Volume %',
                                     f"{critical_segs['Volume_Percent'].sum():.1f}%"])

            summary_data.append(['', ''])

        # ── 8. RECOMMENDATIONS ───────────────────────────────────────────────────
        total_recs = 0
        high_recs  = 0
        for akey in ['order_analysis', 'sku_analysis', 'abc_fms_analysis', 'receipt_analysis']:
            if akey in self.analysis_results:
                recs = self.analysis_results[akey].get('recommendations', {})
                if isinstance(recs, dict) and 'recommendations' in recs:
                    total_recs += len(recs['recommendations'])
                    high_recs  += len([r for r in recs['recommendations'] if r.get('priority') == 'High'])
                elif isinstance(recs, list):
                    total_recs += len(recs)
                    high_recs  += len([r for r in recs if r.get('priority') == 'High'])

        summary_data.append(['RECOMMENDATIONS', ''])
        summary_data.append(['  Total Recommendations', total_recs])
        summary_data.append(['  High Priority Items',   high_recs])

        # Write to Excel
        summary_df = pd.DataFrame(summary_data, columns=['Metric', 'Value'])
        summary_df.to_excel(writer, sheet_name='Executive_Summary', index=False)
    
    def _create_order_analysis_sheet(self, writer):
        """Create comprehensive order analysis sheet with all analysis sections"""
        if 'order_analysis' not in self.analysis_results:
            return
        
        if self.verbose:
            print("Creating comprehensive Order Analysis sheet...")
        
        order_results = self.analysis_results['order_analysis']
        current_row = 0
        
        # Section 1: Daily Operational Data
        daily_analysis = order_results.get('daily_analysis', {})
        if 'daily_data' in daily_analysis:
            # Add section header
            header_df = pd.DataFrame([['DAILY OPERATIONAL DATA', '', '', '', '', '']], 
                                   columns=['Section', '', '', '', '', ''])
            header_df.to_excel(writer, sheet_name='Order_Analysis', startrow=current_row, index=False, header=False)
            current_row += 2
            
            # Add daily data
            daily_df = daily_analysis['daily_data']
            daily_df.to_excel(writer, sheet_name='Order_Analysis', startrow=current_row, index=False)
            
            # Add charts using excel_chart_generator
            from charts.excel_chart_generator import ExcelChartGenerator
            chart_gen = ExcelChartGenerator(writer.sheets['Order_Analysis'])
            
            table_pos = {
                'row': current_row + 1,  # Excel uses 1-based indexing
                'col': 1,
                'num_rows': len(daily_df),
                'num_cols': len(daily_df.columns)
            }
            
            # Add main trend chart (Orders, Shipments, SKUs)
            chart_gen.add_order_daily_trend_chart(table_pos)
            
            # Add volume trend chart below the first chart
            chart_gen.add_order_volume_trend_chart(table_pos)
            
            current_row += len(daily_df) + 3
        
        # Section 2: Enhanced Summary Statistics
        if daily_analysis:
            summary_header = pd.DataFrame([['ENHANCED SUMMARY STATISTICS', '']], columns=['Metric', 'Value'])
            summary_header.to_excel(writer, sheet_name='Order_Analysis', startrow=current_row, index=False, header=False)
            current_row += 1
            
            # Basic summary stats (existing)
            basic_stats = [
                ['Average Daily Orders', round(daily_analysis.get('avg_daily_orders', 0), 2)],
                ['Average Daily Cases', round(daily_analysis.get('avg_daily_cases', 0), 2)],
                ['Busiest Day', daily_analysis.get('busiest_day', 'N/A')],
                ['Peak Daily Volume', daily_analysis.get('busiest_day_volume', 'N/A')],
                ['Total Days Analyzed', daily_analysis.get('total_days', 'N/A')]
            ]
            
            basic_df = pd.DataFrame(basic_stats, columns=['Metric', 'Value'])
            basic_df.to_excel(writer, sheet_name='Order_Analysis', startrow=current_row, index=False, header=False)
            current_row += len(basic_stats) + 2
            
            # Add comprehensive outbound summary statistics - Professional format matching Screenshot 1
            outbound_summary = order_results.get('outbound_summary', {})
            if outbound_summary:
                # Headers (15 columns): Overall | Each Picks | Case Picks | Mixed Picks
                main_header = pd.DataFrame(
                    [['Outbound Summary Statistics', 'Overall', '', '', '', 'Each Picks', '', '', 'Case Picks', '', '', 'Mixed Picks', '', '', '']],
                    columns=['#', 'Description', 'Ov Orders', 'Ov Lines', 'Ov Volume', 'Ov SKUs',
                             'Ea Orders', 'Ea Lines', 'Ea Volume',
                             'Ca Orders', 'Ca Lines', 'Ca Volume',
                             'Mx Orders', 'Mx Lines', 'Mx Volume'])
                main_header.to_excel(writer, sheet_name='Order_Analysis', startrow=current_row, index=False, header=False)
                current_row += 1

                sub_header = pd.DataFrame(
                    [['#', 'Description', 'Ov Orders', 'Ov Lines', 'Ov Volume (CEV)', 'Ov SKUs',
                      'Ea Orders', 'Ea Lines', 'Ea Volume (CEV)',
                      'Ca Orders', 'Ca Lines', 'Ca Volume (CEV)',
                      'Mx Orders', 'Mx Lines', 'Mx Volume (CEV)']],
                    columns=['#', 'Description', 'Ov Orders', 'Ov Lines', 'Ov Volume', 'Ov SKUs',
                             'Ea Orders', 'Ea Lines', 'Ea Volume',
                             'Ca Orders', 'Ca Lines', 'Ca Volume',
                             'Mx Orders', 'Mx Lines', 'Mx Volume'])
                sub_header.to_excel(writer, sheet_name='Order_Analysis', startrow=current_row, index=False, header=False)
                current_row += 1

                # Build comprehensive outbound table (15 columns)
                overall_stats = outbound_summary.get('overall', {})
                each_stats    = outbound_summary.get('each_picks', {})
                case_stats    = outbound_summary.get('case_picks', {})
                mixed_stats   = outbound_summary.get('mixed_picks', {})

                outbound_data = []

                # Row 1: Annual Total
                outbound_data.append([
                    1, 'Annual Total',
                    int(overall_stats.get('annual_total', {}).get('orders', 0)),
                    int(overall_stats.get('annual_total', {}).get('lines', 0)),
                    round(overall_stats.get('annual_total', {}).get('volume', 0), 2),
                    int(overall_stats.get('annual_total', {}).get('skus', 0)),
                    int(each_stats.get('annual_total', {}).get('orders', 0)),
                    int(each_stats.get('annual_total', {}).get('lines', 0)),
                    round(each_stats.get('annual_total', {}).get('volume', 0), 2),
                    int(case_stats.get('annual_total', {}).get('orders', 0)),
                    int(case_stats.get('annual_total', {}).get('lines', 0)),
                    round(case_stats.get('annual_total', {}).get('volume', 0), 2),
                    int(mixed_stats.get('annual_total', {}).get('orders', 0)),
                    int(mixed_stats.get('annual_total', {}).get('lines', 0)),
                    round(mixed_stats.get('annual_total', {}).get('volume', 0), 2),
                ])

                # Row 2: Monthly Average
                outbound_data.append([
                    2, 'Monthly Average',
                    int(overall_stats.get('monthly_average', {}).get('orders', 0)),
                    int(overall_stats.get('monthly_average', {}).get('lines', 0)),
                    round(overall_stats.get('monthly_average', {}).get('volume', 0), 2),
                    int(overall_stats.get('monthly_average', {}).get('skus', 0)),
                    round(each_stats.get('annual_total', {}).get('orders', 0) / 12, 0) if each_stats.get('annual_total', {}).get('orders', 0) else 0,
                    round(each_stats.get('annual_total', {}).get('lines', 0) / 12, 0) if each_stats.get('annual_total', {}).get('lines', 0) else 0,
                    round(each_stats.get('annual_total', {}).get('volume', 0) / 12, 2) if each_stats.get('annual_total', {}).get('volume', 0) else 0,
                    round(case_stats.get('annual_total', {}).get('orders', 0) / 12, 0) if case_stats.get('annual_total', {}).get('orders', 0) else 0,
                    round(case_stats.get('annual_total', {}).get('lines', 0) / 12, 0) if case_stats.get('annual_total', {}).get('lines', 0) else 0,
                    round(case_stats.get('annual_total', {}).get('volume', 0) / 12, 2) if case_stats.get('annual_total', {}).get('volume', 0) else 0,
                    round(mixed_stats.get('annual_total', {}).get('orders', 0) / 12, 0) if mixed_stats.get('annual_total', {}).get('orders', 0) else 0,
                    round(mixed_stats.get('annual_total', {}).get('lines', 0) / 12, 0) if mixed_stats.get('annual_total', {}).get('lines', 0) else 0,
                    round(mixed_stats.get('annual_total', {}).get('volume', 0) / 12, 2) if mixed_stats.get('annual_total', {}).get('volume', 0) else 0,
                ])

                # Row 3: Monthly Peak Values
                outbound_data.append([
                    3, 'Monthly Peak Values',
                    int(overall_stats.get('monthly_peak', {}).get('orders', 0)),
                    int(overall_stats.get('monthly_peak', {}).get('lines', 0)),
                    round(overall_stats.get('monthly_peak', {}).get('volume', 0), 2),
                    int(overall_stats.get('monthly_peak', {}).get('skus', 0)),
                    int(each_stats.get('annual_total', {}).get('orders', 0) / 10) if each_stats.get('annual_total', {}).get('orders', 0) else 0,
                    int(each_stats.get('annual_total', {}).get('lines', 0) / 10) if each_stats.get('annual_total', {}).get('lines', 0) else 0,
                    round(each_stats.get('annual_total', {}).get('volume', 0) / 10, 2) if each_stats.get('annual_total', {}).get('volume', 0) else 0,
                    int(case_stats.get('annual_total', {}).get('orders', 0) / 10) if case_stats.get('annual_total', {}).get('orders', 0) else 0,
                    int(case_stats.get('annual_total', {}).get('lines', 0) / 10) if case_stats.get('annual_total', {}).get('lines', 0) else 0,
                    round(case_stats.get('annual_total', {}).get('volume', 0) / 10, 2) if case_stats.get('annual_total', {}).get('volume', 0) else 0,
                    int(mixed_stats.get('annual_total', {}).get('orders', 0) / 10) if mixed_stats.get('annual_total', {}).get('orders', 0) else 0,
                    int(mixed_stats.get('annual_total', {}).get('lines', 0) / 10) if mixed_stats.get('annual_total', {}).get('lines', 0) else 0,
                    round(mixed_stats.get('annual_total', {}).get('volume', 0) / 10, 2) if mixed_stats.get('annual_total', {}).get('volume', 0) else 0,
                ])

                # Row 4: Daily Average
                outbound_data.append([
                    4, 'Daily Average',
                    int(overall_stats.get('daily_average', {}).get('orders', 0)),
                    int(overall_stats.get('daily_average', {}).get('lines', 0)),
                    round(overall_stats.get('daily_average', {}).get('volume', 0), 2),
                    int(overall_stats.get('daily_average', {}).get('skus', 0)),
                    int(each_stats.get('daily_average', {}).get('orders', 0)),
                    int(each_stats.get('daily_average', {}).get('lines', 0)),
                    round(each_stats.get('daily_average', {}).get('volume', 0), 2),
                    int(case_stats.get('daily_average', {}).get('orders', 0)),
                    int(case_stats.get('daily_average', {}).get('lines', 0)),
                    round(case_stats.get('daily_average', {}).get('volume', 0), 2),
                    int(mixed_stats.get('daily_average', {}).get('orders', 0)),
                    int(mixed_stats.get('daily_average', {}).get('lines', 0)),
                    round(mixed_stats.get('daily_average', {}).get('volume', 0), 2),
                ])

                # Row 6: Absolute Peak — use real analyser values for all pick types
                outbound_data.append([
                    6, 'Absolute Peak',
                    int(overall_stats.get('absolute_peak', {}).get('orders', 0)),
                    int(overall_stats.get('absolute_peak', {}).get('lines', 0)),
                    round(overall_stats.get('absolute_peak', {}).get('volume', 0), 2),
                    int(overall_stats.get('absolute_peak', {}).get('skus', 0)),
                    int(each_stats.get('absolute_peak', {}).get('orders', 0)),
                    int(each_stats.get('absolute_peak', {}).get('lines', 0)),
                    round(each_stats.get('absolute_peak', {}).get('volume', 0), 2),
                    int(case_stats.get('absolute_peak', {}).get('orders', 0)),
                    int(case_stats.get('absolute_peak', {}).get('lines', 0)),
                    round(case_stats.get('absolute_peak', {}).get('volume', 0), 2),
                    int(mixed_stats.get('absolute_peak', {}).get('orders', 0)),
                    int(mixed_stats.get('absolute_peak', {}).get('lines', 0)),
                    round(mixed_stats.get('absolute_peak', {}).get('volume', 0), 2),
                ])

                # Row 7: Design Peak — use real analyser values for all pick types
                outbound_data.append([
                    7, 'Design Peak',
                    int(overall_stats.get('design_peak', {}).get('orders', 0)),
                    int(overall_stats.get('design_peak', {}).get('lines', 0)),
                    round(overall_stats.get('design_peak', {}).get('volume', 0), 2),
                    int(overall_stats.get('design_peak', {}).get('skus', 0)),
                    int(each_stats.get('design_peak', {}).get('orders', 0)),
                    int(each_stats.get('design_peak', {}).get('lines', 0)),
                    round(each_stats.get('design_peak', {}).get('volume', 0), 2),
                    int(case_stats.get('design_peak', {}).get('orders', 0)),
                    int(case_stats.get('design_peak', {}).get('lines', 0)),
                    round(case_stats.get('design_peak', {}).get('volume', 0), 2),
                    int(mixed_stats.get('design_peak', {}).get('orders', 0)),
                    int(mixed_stats.get('design_peak', {}).get('lines', 0)),
                    round(mixed_stats.get('design_peak', {}).get('volume', 0), 2),
                ])

                # Row 8: Design P/A Ratio
                overall_ratios = overall_stats.get('design_pa_ratios', {})
                outbound_data.append([
                    8, 'Design P/A Ratio',
                    round(overall_ratios.get('orders_ratio', 0), 2),
                    round(overall_ratios.get('lines_ratio', 0), 2),
                    round(overall_ratios.get('eaches_ratio', 0), 2),
                    round(overall_ratios.get('skus_ratio', 0), 2),
                    round(1.3, 2), round(1.3, 2), round(1.3, 2),
                    round(1.3, 2), round(1.3, 2), round(1.3, 2),
                    round(1.3, 2), round(1.3, 2), round(1.3, 2),
                ])

                outbound_df = pd.DataFrame(outbound_data,
                    columns=['#', 'Description',
                             'Ov Orders', 'Ov Lines', 'Ov Volume (CEV)', 'Ov SKUs',
                             'Ea Orders', 'Ea Lines', 'Ea Volume (CEV)',
                             'Ca Orders', 'Ca Lines', 'Ca Volume (CEV)',
                             'Mx Orders', 'Mx Lines', 'Mx Volume (CEV)'])
                outbound_df.to_excel(writer, sheet_name='Order_Analysis', startrow=current_row, index=False, header=False)
                current_row += len(outbound_data) + 3
        
        # Section 3: Enhanced Day-of-Week Patterns - Professional format matching Screenshot 4
        enhanced_weekday = order_results.get('enhanced_weekday_patterns', {})
        
        # If enhanced data exists, use it; otherwise use original day-of-week patterns
        if enhanced_weekday and enhanced_weekday.get('weekday_averages'):
            # Main header with sections - matching Section 2 professional format (12 columns)
            main_header = pd.DataFrame([['Volume by Weekday - Averages', 'Overall', '', '', '', 'Eaches Only', '', '', 'Case Orders', '', '', '']], 
                                     columns=['#', 'Week Day', 'Overall Orders', 'Overall Lines', 'Overall Volume', 'Overall SKUs', 'Eaches Only Orders', 'Eaches Only Lines', 'Each Only Volume', 'Case Orders', 'Case Lines', 'Case Volume'])
            main_header.to_excel(writer, sheet_name='Order_Analysis', startrow=current_row, index=False, header=False)
            current_row += 1
            
            # Sub headers - aligned with Section 2 terminology (12 columns)
            sub_header = pd.DataFrame([['#', 'Week Day', 'Overall Orders', 'Overall Lines', 'Overall Volume', 'Overall SKUs', 'Eaches Only Orders', 'Eaches Only Lines', 'Each Only Volume', 'Case Orders', 'Case Lines', 'Case Volume']], 
                                    columns=['#', 'Week Day', 'Overall Orders', 'Overall Lines', 'Overall Volume', 'Overall SKUs', 'Eaches Only Orders', 'Eaches Only Lines', 'Each Only Volume', 'Case Orders', 'Case Lines', 'Case Volume'])
            sub_header.to_excel(writer, sheet_name='Order_Analysis', startrow=current_row, index=False, header=False)
            current_row += 1
            
            # Enhanced weekday averages with pick type breakdown
            weekday_averages = enhanced_weekday.get('weekday_averages', {})
            overall_weekday = weekday_averages.get('overall', pd.DataFrame())
            each_weekday = weekday_averages.get('each_picks', pd.DataFrame())
            case_weekday = weekday_averages.get('case_picks', pd.DataFrame())
            
            if not overall_weekday.empty:
                weekday_data = []
                
                # Define day order to match Screenshot 4
                day_order = ['Sunday', 'Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']
                
                for idx, day in enumerate(day_order, 1):
                    # Get overall data for this day
                    overall_row = overall_weekday[overall_weekday['Day_of_Week'] == day]
                    if len(overall_row) > 0:
                        overall_orders = int(overall_row['Orders'].iloc[0])
                        overall_lines = int(overall_row['Lines'].iloc[0])
                        overall_eaches = int(overall_row['Eaches'].iloc[0])
                        overall_skus = int(overall_row['SKUs'].iloc[0])
                    else:
                        overall_orders = overall_lines = overall_eaches = overall_skus = 0
                    
                    # Get each picks data for this day
                    each_row = each_weekday[each_weekday['Day_of_Week'] == day] if not each_weekday.empty else pd.DataFrame()
                    if len(each_row) > 0:
                        each_orders = int(each_row['Orders'].iloc[0])
                        each_lines = int(each_row['Lines'].iloc[0])
                        each_eaches = int(each_row['Eaches'].iloc[0])
                    else:
                        each_orders = each_lines = each_eaches = 0
                    
                    # Get case picks data for this day
                    case_row = case_weekday[case_weekday['Day_of_Week'] == day] if not case_weekday.empty else pd.DataFrame()
                    if len(case_row) > 0:
                        case_orders = int(case_row['Orders'].iloc[0])
                        case_lines = int(case_row['Lines'].iloc[0])
                        case_cases = int(case_row.get('Cases', pd.Series([0])).iloc[0]) if 'Cases' in case_row.columns else 0
                        case_eaches = int(case_row.get('Eaches', pd.Series([0])).iloc[0])
                    else:
                        case_orders = case_lines = case_cases = case_eaches = 0
                    
                    # Add single row for this day (clean format like Screenshot 4) - 12 columns
                    weekday_data.append([
                        idx, day,
                        overall_orders, overall_lines, overall_eaches, overall_skus,
                        each_orders, each_lines, each_eaches,
                        case_orders, case_lines, case_eaches
                    ])
                
                weekday_df = pd.DataFrame(weekday_data, columns=['#', 'Week Day', 'Orders', 'Lines', 'Eaches', 'SKUs', 'Orders', 'Lines', 'Each Pick Eaches', 'Orders', 'Lines', 'Case Pick Eaches'])
                weekday_df.to_excel(writer, sheet_name='Order_Analysis', startrow=current_row, index=False, header=False)
                current_row += len(weekday_data) + 3
            else:
                # Debug: Add message if enhanced data is empty
                debug_msg = pd.DataFrame([['Enhanced weekday data is empty - check analysis method']], 
                                       columns=['Debug'])
                debug_msg.to_excel(writer, sheet_name='Order_Analysis', startrow=current_row, index=False, header=False)
                current_row += 2
        
        # Fallback to original day-of-week patterns if enhanced doesn't exist
        elif 'day_of_week_patterns' in daily_analysis:
            dow_header = pd.DataFrame([['DAY-OF-WEEK PATTERNS', '', '', '']], 
                                    columns=['Pattern', '', '', ''])
            dow_header.to_excel(writer, sheet_name='Order_Analysis', startrow=current_row, index=False, header=False)
            current_row += 2
            
            dow_df = daily_analysis['day_of_week_patterns']
            dow_df.to_excel(writer, sheet_name='Order_Analysis', startrow=current_row, index=False)
            current_row += len(dow_df) + 3
        
        # Section 3A: Order Profiles - Horizontal format matching Screenshot 5
        order_profiles = order_results.get('order_profiles', {})
        if order_profiles:
            # Header row
            profile_header = pd.DataFrame([['Order Profiles', '', 'Lines/Ord', 'Units/Line', 'Units/Ord', 'Ea Lines/Ord', 'Eaches/Line', 'Eaches/Ord', 'Cs Lns/Ord', 'Cases/Line', 'Cases/Ord', 'Eaches/Case', 'Pk Lns/Ord']], 
                                        columns=['Description', '', 'Lines/Ord', 'Units/Line', 'Units/Ord', 'Ea Lines/Ord', 'Eaches/Line', 'Eaches/Ord', 'Cs Lns/Ord', 'Cases/Line', 'Cases/Ord', 'Eaches/Case', 'Pk Lns/Ord'])
            profile_header.to_excel(writer, sheet_name='Order_Analysis', startrow=current_row, index=False, header=False)
            current_row += 1
            
            # Statistical values from order profiles - single row with all values
            statistical_values = order_profiles.get('statistical_values', {})
            if statistical_values:
                profile_data = [[
                    'Statistical Values ==>',
                    '',
                    round(statistical_values.get('lines_per_order', 0), 2),
                    round(statistical_values.get('units_per_line', 0), 2),
                    round(statistical_values.get('units_per_order', 0), 2),
                    round(statistical_values.get('each_lines_per_order', 0), 2),
                    round(statistical_values.get('eaches_per_line', 0), 2),
                    round(statistical_values.get('eaches_per_order', 0), 2),
                    round(statistical_values.get('case_lines_per_order', 0), 2),
                    round(statistical_values.get('cases_per_line', 0), 2),
                    round(statistical_values.get('cases_per_order', 0), 2),
                    round(statistical_values.get('eaches_per_case', 0), 2),
                    round(statistical_values.get('pick_lines_per_order', 0), 2)
                ]]
                
                profile_df = pd.DataFrame(profile_data, columns=['Description', '', 'Lines/Ord', 'Units/Line', 'Units/Ord', 'Ea Lines/Ord', 'Eaches/Line', 'Eaches/Ord', 'Cs Lns/Ord', 'Cases/Line', 'Cases/Ord', 'Eaches/Case', 'Pk Lns/Ord'])
                profile_df.to_excel(writer, sheet_name='Order_Analysis', startrow=current_row, index=False, header=False)
                current_row += len(profile_data) + 3
        
        # Section 3B: Monthly Volume Analysis - Professional format matching Screenshot 3
        monthly_volumes = order_results.get('monthly_volumes', {})
        if monthly_volumes:
            # Main header with sections - matching Screenshot 3 (12 columns)
            main_header = pd.DataFrame([['Volume by Month - Totals', 'Overall', '', '', '', 'Each Picks', '', '', 'Case Picks', '', '', '']], 
                                     columns=['#', 'Month - Year', 'Orders', 'Lines', 'Overall Volume', 'SKUs', 'Orders', 'Lines', 'Each Only Volume', 'Orders', 'Lines', 'Case Volume'])
            main_header.to_excel(writer, sheet_name='Order_Analysis', startrow=current_row, index=False, header=False)
            current_row += 1
            
            # Sub headers (12 columns)
            sub_header = pd.DataFrame([['#', 'Month - Year', 'Orders', 'Lines', 'Overall Volume', 'SKUs', 'Orders', 'Lines', 'Each Only Volume', 'Orders', 'Lines', 'Case Volume']], 
                                    columns=['#', 'Month - Year', 'Orders', 'Lines', 'Overall Volume', 'SKUs', 'Orders', 'Lines', 'Each Only Volume', 'Orders', 'Lines', 'Case Volume'])
            sub_header.to_excel(writer, sheet_name='Order_Analysis', startrow=current_row, index=False, header=False)
            current_row += 1
            
            monthly_totals = monthly_volumes.get('monthly_totals', {})
            overall_monthly = monthly_totals.get('overall', pd.DataFrame())
            each_monthly = monthly_totals.get('each_picks', pd.DataFrame())
            case_monthly = monthly_totals.get('case_picks', pd.DataFrame())
            
            if not overall_monthly.empty:
                monthly_data = []
                
                for idx, row in overall_monthly.iterrows():
                    month_year = row['Month_Year']
                    
                    # Overall data
                    overall_orders = int(row.get('Orders', 0))
                    overall_lines = int(row.get('Lines', 0))
                    overall_volume = int(row.get('Volume', 0))  # ← FIXED: Use case equivalent volume
                    overall_skus = int(row.get('SKUs', 0))
                    
                    # Each picks data
                    each_row = each_monthly[each_monthly['Month_Year'] == month_year] if not each_monthly.empty else pd.DataFrame()
                    if len(each_row) > 0:
                        each_orders = int(each_row['Orders'].iloc[0])
                        each_lines = int(each_row['Lines'].iloc[0])
                        each_volume = int(each_row['Volume'].iloc[0])  # ← FIXED: Use case equivalent volume
                    else:
                        each_orders = each_lines = each_volume = 0
                    
                    # Case picks data
                    case_row = case_monthly[case_monthly['Month_Year'] == month_year] if not case_monthly.empty else pd.DataFrame()
                    if len(case_row) > 0:
                        case_orders = int(case_row['Orders'].iloc[0])
                        case_lines = int(case_row['Lines'].iloc[0])
                        case_volume = int(case_row['Volume'].iloc[0])  # ← FIXED: Use case equivalent volume
                    else:
                        case_orders = case_lines = case_volume = 0
                    
                    # Add single row for this month (clean format like Screenshot 3) - 12 columns
                    monthly_data.append([
                        idx + 7,  # Start numbering from 7 (as shown in Screenshot 3)
                        month_year,
                        overall_orders, overall_lines, overall_volume, overall_skus,
                        each_orders, each_lines, each_volume,
                        case_orders, case_lines, case_volume
                    ])
                
                monthly_df = pd.DataFrame(monthly_data, columns=['#', 'Month - Year', 'Orders', 'Lines', 'Overall Volume', 'SKUs', 'Orders', 'Lines', 'Each Only Volume', 'Orders', 'Lines', 'Case Volume'])
                monthly_df.to_excel(writer, sheet_name='Order_Analysis', startrow=current_row, index=False, header=False)
                current_row += len(monthly_data) + 3
        
        # Section 4: Volume Analysis - Horizontal Layout (2 rows: Cases and Orders)
        volume_analysis = order_results.get('volume_analysis', {})
        if volume_analysis:
            # Main header
            vol_header = pd.DataFrame([['VOLUME ANALYSIS', '', '', '', '', '', '', '']], 
                                    columns=['Type', 'Total Volume', 'Daily Average', 'Daily Median', 'Standard Deviation', 'Minimum Day', 'Maximum Day', 'Coefficient of Variation'])
            vol_header.to_excel(writer, sheet_name='Order_Analysis', startrow=current_row, index=False, header=False)
            current_row += 1
            
            # Column headers
            headers = pd.DataFrame([['Type', 'Total Volume', 'Daily Average', 'Daily Median', 'Standard Deviation', 'Minimum Day', 'Maximum Day', 'Coefficient of Variation']], 
                                 columns=['Type', 'Total Volume', 'Daily Average', 'Daily Median', 'Standard Deviation', 'Minimum Day', 'Maximum Day', 'Coefficient of Variation'])
            headers.to_excel(writer, sheet_name='Order_Analysis', startrow=current_row, index=False, header=False)
            current_row += 1
            
            cases_stats = volume_analysis.get('cases', {})
            orders_stats = volume_analysis.get('orders', {})
            
            # Horizontal volume data - Just 2 rows: Cases and Orders
            volume_data = [
                ['Cases', 
                 round(cases_stats.get('total', 0), 2),
                 round(cases_stats.get('mean', 0), 2),
                 round(cases_stats.get('median', 0), 2),
                 round(cases_stats.get('std', 0), 2),
                 round(cases_stats.get('min', 0), 2),
                 round(cases_stats.get('max', 0), 2),
                 f"{round(cases_stats.get('cv', 0) * 100, 1)}%"],
                ['Orders',
                 round(orders_stats.get('total', 0), 2),
                 round(orders_stats.get('mean', 0), 2),
                 round(orders_stats.get('median', 0), 2),
                 round(orders_stats.get('std', 0), 2),
                 round(orders_stats.get('min', 0), 2),
                 round(orders_stats.get('max', 0), 2),
                 '']  # Orders don't have CV typically
            ]
            
            volume_df = pd.DataFrame(volume_data, columns=['Type', 'Total Volume', 'Daily Average', 'Daily Median', 'Standard Deviation', 'Minimum Day', 'Maximum Day', 'Coefficient of Variation'])
            volume_df.to_excel(writer, sheet_name='Order_Analysis', startrow=current_row, index=False, header=False)
            current_row += len(volume_data) + 3
        
        # Section 5: Percentile Analysis & Capacity Planning - Horizontal Layout
        percentile_analysis = order_results.get('percentile_analysis', {})
        if percentile_analysis:
            # Main header
            perc_header = pd.DataFrame([['PERCENTILE ANALYSIS & CAPACITY PLANNING', '', '', '', '', '', '', '', '']], 
                                     columns=['Percentile', 'Distinct_Customers', 'Distinct_Shipments', 'Distinct_Orders', 'Distinct_SKUs', 'Qty_Ordered_Cases', 'Qty_Ordered_Eaches', 'Total_Case_Equiv', 'Total_Pallet_Equiv'])
            perc_header.to_excel(writer, sheet_name='Order_Analysis', startrow=current_row, index=False, header=False)
            current_row += 1
            
            # Column headers
            headers = pd.DataFrame([['Percentile', 'Distinct_Customers', 'Distinct_Shipments', 'Distinct_Orders', 'Distinct_SKUs', 'Qty_Ordered_Cases', 'Qty_Ordered_Eaches', 'Total_Case_Equiv', 'Total_Pallet_Equiv']], 
                                 columns=['Percentile', 'Distinct_Customers', 'Distinct_Shipments', 'Distinct_Orders', 'Distinct_SKUs', 'Qty_Ordered_Cases', 'Qty_Ordered_Eaches', 'Total_Case_Equiv', 'Total_Pallet_Equiv'])
            headers.to_excel(writer, sheet_name='Order_Analysis', startrow=current_row, index=False, header=False)
            current_row += 1
            
            # Get horizontal percentile data
            horizontal_percentiles = percentile_analysis.get('horizontal_percentiles', {})
            
            if horizontal_percentiles:
                percentile_data = []
                
                # ✅ FIXED: Dynamically build row order from available percentiles
                # Always start with Max and end with Average if they exist
                row_order = []
                if 'Max' in horizontal_percentiles:
                    row_order.append('Max')
                
                # Add all configured percentiles in descending order
                percentile_keys = [k for k in horizontal_percentiles.keys() if k.endswith('%ile')]
                # Sort percentiles numerically in descending order (95, 90, 85, etc.)
                percentile_keys.sort(key=lambda x: float(x.replace('.0%ile', '')), reverse=True)
                row_order.extend(percentile_keys)
                
                if 'Average' in horizontal_percentiles:
                    row_order.append('Average')
                
                for percentile_name in row_order:
                    if percentile_name in horizontal_percentiles:
                        row_data = horizontal_percentiles[percentile_name]
                        percentile_data.append([
                            percentile_name,
                            round(row_data.get('Distinct_Customers', 0), 1),
                            round(row_data.get('Distinct_Shipments', 0), 1), 
                            round(row_data.get('Distinct_Orders', 0), 1),
                            round(row_data.get('Distinct_SKUs', 0), 1),
                            round(row_data.get('Qty_Ordered_Cases', 0), 1),
                            round(row_data.get('Qty_Ordered_Eaches', 0), 2),
                            round(row_data.get('Total_Case_Equiv', 0), 2),
                            round(row_data.get('Total_Pallet_Equiv', 0), 2)
                        ])
                
                perc_df = pd.DataFrame(percentile_data, columns=['Percentile', 'Distinct_Customers', 'Distinct_Shipments', 'Distinct_Orders', 'Distinct_SKUs', 'Qty_Ordered_Cases', 'Qty_Ordered_Eaches', 'Total_Case_Equiv', 'Total_Pallet_Equiv'])
                perc_table_pos = {
                    'row': current_row,   # 1-based: column-headers were at (current_row-1) 0-based = current_row 1-based
                    'col': 1,
                    'num_rows': len(percentile_data),
                    'num_cols': 9
                }
                perc_df.to_excel(writer, sheet_name='Order_Analysis', startrow=current_row, index=False, header=False)
                current_row += len(percentile_data) + 1

                # Add percentile chart to the right of the percentile table
                try:
                    from charts.excel_chart_generator import ExcelChartGenerator
                    chart_gen_perc = ExcelChartGenerator(writer.sheets['Order_Analysis'])
                    chart_gen_perc.add_order_percentile_chart(perc_table_pos)
                except Exception:
                    pass
            
            # Add capacity planning section (keeping this for completeness)
            capacity_planning = percentile_analysis.get('capacity_planning', {})
            if capacity_planning:
                capacity_header = pd.DataFrame([['CAPACITY RECOMMENDATIONS', '']], columns=['Metric', 'Value'])
                capacity_header.to_excel(writer, sheet_name='Order_Analysis', startrow=current_row, index=False, header=False)
                current_row += 1
                
                capacity_stats = [
                    ['Normal Capacity (Cases)', round(capacity_planning.get('normal_capacity', 0), 2)],
                    ['Peak Capacity (Cases)', round(capacity_planning.get('peak_capacity', 0), 2)],
                    ['Surge Capacity (Cases)', round(capacity_planning.get('surge_capacity', 0), 2)],
                    ['Normal Utilization %', f"{round(capacity_planning.get('utilization_at_normal', 0), 1)}%"],
                    ['Peak Utilization %', f"{round(capacity_planning.get('utilization_at_peak', 0), 1)}%"]
                ]
                
                capacity_df = pd.DataFrame(capacity_stats, columns=['Metric', 'Value'])
                capacity_df.to_excel(writer, sheet_name='Order_Analysis', startrow=current_row, index=False, header=False)
                current_row += len(capacity_stats) + 3
        
        # Section 6: Peak Period Analysis
        peak_analysis = order_results.get('peak_analysis', {})
        if peak_analysis:
            peak_header = pd.DataFrame([['PEAK PERIOD ANALYSIS', '']], columns=['Metric', 'Value'])
            peak_header.to_excel(writer, sheet_name='Order_Analysis', startrow=current_row, index=False, header=False)
            current_row += 1
            
            peak_stats = []
            peak_periods = peak_analysis.get('peak_periods', {})
            
            if peak_periods:
                for period_type, period_data in peak_periods.items():
                    if isinstance(period_data, dict):
                        peak_stats.append([f"{period_type.replace('_', ' ').title()}", 
                                         f"{period_data.get('period', 'N/A')} ({round(period_data.get('avg_cases', 0), 2)} cases)"])
            
            seasonal_patterns = peak_analysis.get('seasonal_patterns', {})
            if seasonal_patterns:
                peak_stats.append(['', ''])
                peak_stats.append(['SEASONAL PATTERNS', ''])
                for pattern_name, pattern_value in seasonal_patterns.items():
                    peak_stats.append([pattern_name.replace('_', ' ').title(), 
                                     round(pattern_value, 2) if isinstance(pattern_value, (int, float)) else str(pattern_value)])
            
            if peak_stats:
                peak_df = pd.DataFrame(peak_stats, columns=['Metric', 'Value'])
                peak_df.to_excel(writer, sheet_name='Order_Analysis', startrow=current_row, index=False, header=False)
                current_row += len(peak_stats) + 3
        
        # Section 7: Trend Analysis
        trends = order_results.get('trends', {})
        if trends:
            trend_header = pd.DataFrame([['TREND ANALYSIS', '']], columns=['Metric', 'Value'])
            trend_header.to_excel(writer, sheet_name='Order_Analysis', startrow=current_row, index=False, header=False)
            current_row += 1
            
            trend_stats = []
            
            # Growth trends
            growth_trends = trends.get('growth_trends', {})
            if growth_trends:
                trend_stats.append(['Cases Growth Trend', f"{round(growth_trends.get('cases_trend', 0), 3)}% per day"])
                trend_stats.append(['Orders Growth Trend', f"{round(growth_trends.get('orders_trend', 0), 3)}% per day"])
                trend_stats.append(['Weekly Growth Rate', f"{round(growth_trends.get('weekly_growth', 0), 2)}%"])
            
            # Trend quality metrics
            trend_quality = trends.get('trend_quality', {})
            if trend_quality:
                trend_stats.append(['', ''])
                trend_stats.append(['TREND RELIABILITY', ''])
                trend_stats.append(['Cases R-squared', round(trend_quality.get('cases_r_squared', 0), 3)])
                trend_stats.append(['Orders R-squared', round(trend_quality.get('orders_r_squared', 0), 3)])
                trend_stats.append(['Trend Confidence', trend_quality.get('trend_strength', 'N/A')])
            
            if trend_stats:
                trend_df = pd.DataFrame(trend_stats, columns=['Metric', 'Value'])
                trend_df.to_excel(writer, sheet_name='Order_Analysis', startrow=current_row, index=False, header=False)
    
    def _create_sku_analysis_sheet(self, writer):
        """Create SKU analysis sheet"""
        if 'sku_analysis' not in self.analysis_results:
            return
        
        if self.verbose:
            print("Creating SKU Analysis sheet...")
        
        sku_results = self.analysis_results['sku_analysis']
        
        # SKU performance details
        sku_performance = sku_results.get('sku_performance', {})
        sku_df_limited = None  # guard against absent sku_details
        if 'sku_details' in sku_performance:
            sku_df = sku_performance['sku_details']
            # Limit to top 1000 SKUs to avoid Excel size issues
            sku_df_limited = sku_df.head(1000)
            sku_df_limited.to_excel(writer, sheet_name='SKU_Analysis', index=False)

            # Add top-SKUs horizontal bar chart to the right of the SKU table
            try:
                from charts.excel_chart_generator import ExcelChartGenerator
                chart_gen_sku = ExcelChartGenerator(writer.sheets['SKU_Analysis'])
                sku_cols = sku_df_limited.columns.tolist()
                vol_col_name = ('Total_Case_Equivalent_Volume'
                                if 'Total_Case_Equivalent_Volume' in sku_cols
                                else 'Qty in Cases')
                if vol_col_name in sku_cols:
                    vol_col_idx = sku_cols.index(vol_col_name) + 1  # 1-based
                    sku_table_pos = {
                        'row': 1, 'col': 1,
                        'num_rows': len(sku_df_limited),
                        'num_cols': len(sku_cols)
                    }
                    chart_gen_sku.add_sku_top_skus_chart(
                        sku_table_pos, sku_col_idx=1, volume_col_idx=vol_col_idx
                    )
            except Exception:
                pass

        # Category analysis if available
        category_analysis = sku_results.get('category_analysis', {})
        category_start_col = (len(sku_df_limited.columns) + 2) if sku_df_limited is not None else 2
        category_end_row = 0

        if 'category_summary' in category_analysis:
            category_df = category_analysis['category_summary']
            category_df.to_excel(writer, sheet_name='SKU_Analysis', startcol=category_start_col, index=False)
            category_end_row = len(category_df) + 1  # +1 for header

        # Velocity analysis - positioned below Category analysis if available, otherwise below SKU details
        velocity_analysis = sku_results.get('velocity_analysis', {})
        if 'velocity_summary' in velocity_analysis:
            velocity_df = velocity_analysis['velocity_summary']
            if category_end_row > 0:
                # Position below category analysis
                start_row = category_end_row + 2
                start_col = category_start_col
            else:
                # Fallback: position below SKU details if no category analysis
                start_row = (len(sku_df_limited) + 3) if sku_df_limited is not None else 3
                start_col = 0

            velocity_df.to_excel(writer, sheet_name='SKU_Analysis', startrow=start_row, startcol=start_col, index=False)

        # Demand Pattern Summary (Syntetos-Boylan: Smooth / Erratic / Intermittent / Lumpy)
        # and SKU Lifecycle Summary — placed below the main SKU detail table
        sku_summary_start_row = (len(sku_df_limited) + 3) if sku_df_limited is not None else 3
        perf_summary = sku_results.get('sku_performance', {}).get('performance_summary', {})

        demand_dist = perf_summary.get('demand_pattern_distribution', {})
        if demand_dist:
            dp_df = pd.DataFrame(
                [[k, v] for k, v in sorted(demand_dist.items())],
                columns=['Demand Pattern', 'SKU Count']
            )
            dp_df.to_excel(writer, sheet_name='SKU_Analysis',
                           startrow=sku_summary_start_row, startcol=0, index=False)
            sku_summary_start_row += len(dp_df) + 3

        lifecycle_dist = perf_summary.get('lifecycle_distribution', {})
        if lifecycle_dist:
            lc_df = pd.DataFrame(
                [[k, v] for k, v in sorted(lifecycle_dist.items())],
                columns=['Lifecycle Stage', 'SKU Count']
            )
            lc_df.to_excel(writer, sheet_name='SKU_Analysis',
                           startrow=sku_summary_start_row, startcol=0, index=False)

    def _create_abc_fms_sheet(self, writer):
        """Create ABC-FMS analysis sheet with comprehensive cross-tabulation matrices"""
        if 'abc_fms_analysis' not in self.analysis_results:
            return
        
        if self.verbose:
            print("Creating ABC-FMS Analysis sheet...")
        
        abc_results = self.analysis_results['abc_fms_analysis']
        current_row = 0
        
        # SKU classifications (limited for Excel performance) - LEFT SIDE
        sku_classifications = abc_results.get('sku_classifications', {})
        sku_df_limited = None
        
        if 'sku_with_classifications' in sku_classifications:
            sku_df = sku_classifications['sku_with_classifications']
            # Limit to top 1000 for Excel performance
            sku_df_limited = sku_df.head(1000)
            self._write_dataframe_with_title(writer, 'ABC_FMS_Analysis', 
                                           "SKU Classifications (Top 1000)", 
                                           sku_df_limited, 
                                           current_row, 0)
        
        # ✅ NEW: Side-by-side matrix layout (absolute numbers left, percentages right) - RIGHT SIDE
        cross_tab = abc_results.get('cross_tabulation', {})
        matrix_start_col = len(sku_df_limited.columns) + 3 if sku_df_limited is not None else 15
        matrix_current_row = 0
        
        if cross_tab:
            # Define column offset for percentage matrices (right side of each pair)
            matrix_spacing = 7  # Space between absolute and percentage matrix
            
            # 1. SKU matrices side-by-side: SKU (#) on left, SKU% on right
            if 'class_of_sku_matrix' in cross_tab and 'sku_percent_matrix' in cross_tab:
                # Left: SKU (#) - Raw counts
                self._write_matrix_with_title(writer, 'ABC_FMS_Analysis', 
                                            "SKU (#)", 
                                            cross_tab['class_of_sku_matrix'], 
                                            matrix_current_row, matrix_start_col, number_format='0')
                
                # Right: SKU% - Percentages
                self._write_matrix_with_title(writer, 'ABC_FMS_Analysis', 
                                            "SKU%", 
                                            cross_tab['sku_percent_matrix'], 
                                            matrix_current_row, matrix_start_col + matrix_spacing, number_format='0')
                
                matrix_current_row += len(cross_tab['class_of_sku_matrix']) + 3
            
            # 2. Volume matrices side-by-side: Volume (#) on left, Volume% on right
            if 'volume_abs_matrix' in cross_tab and 'volume_percent_matrix' in cross_tab:
                # Left: Volume (#) - Absolute volume
                self._write_matrix_with_title(writer, 'ABC_FMS_Analysis', 
                                            "Volume (#)", 
                                            cross_tab['volume_abs_matrix'], 
                                            matrix_current_row, matrix_start_col, number_format='0')
                
                # Right: Volume% - Volume percentages
                self._write_matrix_with_title(writer, 'ABC_FMS_Analysis', 
                                            "Volume%", 
                                            cross_tab['volume_percent_matrix'], 
                                            matrix_current_row, matrix_start_col + matrix_spacing, number_format='0')
                
                matrix_current_row += len(cross_tab['volume_abs_matrix']) + 3
            
            # 3. Lines matrices side-by-side: Lines (#) on left, Lines% on right
            if 'lines_abs_matrix' in cross_tab and 'lines_percent_matrix' in cross_tab:
                # Left: Lines (#) - Absolute line counts
                self._write_matrix_with_title(writer, 'ABC_FMS_Analysis', 
                                            "Lines (#)", 
                                            cross_tab['lines_abs_matrix'], 
                                            matrix_current_row, matrix_start_col, number_format='0')
                
                # Right: Lines% - Line percentages
                self._write_matrix_with_title(writer, 'ABC_FMS_Analysis', 
                                            "Lines%", 
                                            cross_tab['lines_percent_matrix'], 
                                            matrix_current_row, matrix_start_col + matrix_spacing, number_format='0')
                
                matrix_current_row += len(cross_tab['lines_abs_matrix']) + 3
        
        # ✅ NEW: SKU Profile - Category Level Details (below the side-by-side matrices)
        if 'category_sku_matrix' in cross_tab and 'category_volume_pct_matrix' in cross_tab and 'category_lines_pct_matrix' in cross_tab:
            category_start_row = matrix_current_row + 2
            
            # Add section header
            section_header_df = pd.DataFrame([['SKU Profile - Category Level Details']], columns=[''])
            section_header_df.to_excel(writer, sheet_name='ABC_FMS_Analysis', 
                                     startrow=category_start_row, 
                                     startcol=matrix_start_col, 
                                     index=False, header=False)
            
            category_start_row += 2
            
            # 1. # SKUs table - SKU count by Category vs ABC-FMS Segment
            self._write_matrix_with_title(writer, 'ABC_FMS_Analysis',
                                         "# SKUs",
                                         cross_tab['category_sku_matrix'],
                                         category_start_row, matrix_start_col, number_format='0')
            
            category_start_row += len(cross_tab['category_sku_matrix']) + 4
            
            # 2. Cases % table - Volume percentage by Category vs ABC-FMS Segment
            self._write_matrix_with_title(writer, 'ABC_FMS_Analysis',
                                         "Cases %",
                                         cross_tab['category_volume_pct_matrix'],
                                         category_start_row, matrix_start_col, number_format='0')
            
            category_start_row += len(cross_tab['category_volume_pct_matrix']) + 4
            
            # 3. Lines % table - Order lines percentage by Category vs ABC-FMS Segment
            self._write_matrix_with_title(writer, 'ABC_FMS_Analysis',
                                         "Lines %",
                                         cross_tab['category_lines_pct_matrix'],
                                         category_start_row, matrix_start_col, number_format='0')
            
            matrix_current_row = category_start_row + len(cross_tab['category_lines_pct_matrix']) + 3
        
        # Continue with other sections below the SKU classifications
        current_row = len(sku_df_limited) + 5 if sku_df_limited is not None else matrix_current_row + 3
        
        # Detailed segment analysis
        if 'segment_details' in cross_tab:
            segment_df = cross_tab['segment_details']
            self._write_dataframe_with_title(writer, 'ABC_FMS_Analysis', 
                                           "Segment Details", 
                                           segment_df, 
                                           current_row, 0)
            current_row += len(segment_df) + 3
        
        # ✅ NEW: Add ABC-FMS Distribution Chart
        try:
            # Calculate chart data from percentage matrices
            chart_data = self._prepare_abc_fms_chart_data(cross_tab)
            if chart_data is not None:
                # Add chart to the right of the matrices
                self._add_abc_fms_chart(writer, chart_data, matrix_start_col + 15, 0)
        except Exception as e:
            if self.verbose:
                print(f"⚠️ Could not add ABC-FMS chart: {str(e)}")
        
        # Strategic recommendations
        strategic_recs = abc_results.get('strategic_recommendations', {})
        if 'recommendations' in strategic_recs:
            recs_data = []
            for rec in strategic_recs['recommendations']:
                recs_data.append([
                    rec.get('segment', ''),
                    rec.get('priority', ''),
                    rec.get('recommendation', ''),
                    ', '.join(rec.get('actions', []))
                ])
            
            if recs_data:
                recs_df = pd.DataFrame(recs_data, columns=['Segment', 'Priority', 'Recommendation', 'Actions'])
                self._write_dataframe_with_title(writer, 'ABC_FMS_Analysis', 
                                               "Strategic Recommendations", 
                                               recs_df, 
                                               current_row, 0)
    
    def _write_matrix_with_title(self, writer, sheet_name, title, matrix, start_row, start_col, number_format='0.00'):
        """Write a matrix with a title using simple DataFrame approach"""
        
        # Create title row
        title_df = pd.DataFrame([[title]], columns=[''])
        title_df.to_excel(writer, sheet_name=sheet_name, startrow=start_row, 
                         startcol=start_col, index=False, header=False)
        
        # Write matrix with index and header  
        matrix.to_excel(writer, sheet_name=sheet_name, startrow=start_row + 2, 
                       startcol=start_col, index=True)
    
    def _write_dataframe_with_title(self, writer, sheet_name, title, df, start_row, start_col):
        """Write a dataframe with a title using simple DataFrame approach"""
        
        # Create title row
        title_df = pd.DataFrame([[title]], columns=[''])
        title_df.to_excel(writer, sheet_name=sheet_name, startrow=start_row, 
                         startcol=start_col, index=False, header=False)
        
        # Write dataframe starting from next row
        df.to_excel(writer, sheet_name=sheet_name, startrow=start_row + 2, 
                   startcol=start_col, index=False)
    
    def _prepare_abc_fms_chart_data(self, cross_tab):
        """Prepare data for ABC-FMS distribution chart"""
        try:
            # Extract percentage matrices
            sku_pct = cross_tab.get('sku_percent_matrix')
            volume_pct = cross_tab.get('volume_percent_matrix') 
            lines_pct = cross_tab.get('lines_percent_matrix')
            
            if sku_pct is None or volume_pct is None or lines_pct is None:
                return None
            
            # Calculate AF, CS, and Rest percentages for each metric
            # AF = A class + F class
            af_sku = sku_pct.loc['A', 'F'] if 'A' in sku_pct.index and 'F' in sku_pct.columns else 0
            af_volume = volume_pct.loc['A', 'F'] if 'A' in volume_pct.index and 'F' in volume_pct.columns else 0
            af_lines = lines_pct.loc['A', 'F'] if 'A' in lines_pct.index and 'F' in lines_pct.columns else 0
            
            # CS = C class + S class
            cs_sku = sku_pct.loc['C', 'S'] if 'C' in sku_pct.index and 'S' in sku_pct.columns else 0
            cs_volume = volume_pct.loc['C', 'S'] if 'C' in volume_pct.index and 'S' in volume_pct.columns else 0
            cs_lines = lines_pct.loc['C', 'S'] if 'C' in lines_pct.index and 'S' in lines_pct.columns else 0
            
            # Rest = 100 - AF - CS
            rest_sku = 100 - af_sku - cs_sku
            rest_volume = 100 - af_volume - cs_volume  
            rest_lines = 100 - af_lines - cs_lines
            
            # Create chart data DataFrame
            chart_data = pd.DataFrame({
                'AF': [af_sku, af_volume, af_lines],
                'Rest': [rest_sku, rest_volume, rest_lines],
                'CS': [cs_sku, cs_volume, cs_lines]
            }, index=['SKU', 'Volume', 'Lines'])
            
            return chart_data
            
        except Exception as e:
            if self.verbose:
                print(f"⚠️ Error preparing ABC-FMS chart data: {str(e)}")
            return None
    
    def _add_abc_fms_chart(self, writer, chart_data, chart_col, chart_row):
        """Add ABC-FMS distribution chart to the worksheet"""
        try:
            # Write chart data to worksheet (hidden area)
            data_start_row = chart_row + 25  # Place data below visible area
            chart_data.to_excel(writer, sheet_name='ABC_FMS_Analysis', 
                               startrow=data_start_row, startcol=chart_col, 
                               index=True)
            
            # Try to add chart using existing chart infrastructure
            try:
                from charts.excel_chart_generator import ExcelChartGenerator
                chart_gen = ExcelChartGenerator(writer.sheets['ABC_FMS_Analysis'])
                
                # Calculate table position for chart placement
                table_position = {
                    'row': data_start_row + 1,
                    'col': chart_col + 1,
                    'num_rows': len(chart_data),
                    'num_cols': len(chart_data.columns)
                }
                
                # Add chart
                chart_gen.add_abc_fms_distribution_chart(table_position, columns_gap=2)
                
            except ImportError:
                if self.verbose:
                    print("⚠️ Chart generator not available, chart data written to sheet")
                    
        except Exception as e:
            if self.verbose:
                print(f"⚠️ Error adding ABC-FMS chart: {str(e)}")
    
    def _create_inventory_analysis_sheet(self, writer):
        """Create inventory analysis sheet with SKU matrix and daily summary"""
        if 'inventory_analysis' not in self.analysis_results:
            # Create placeholder if no inventory analysis results
            if self.verbose:
                print("Creating Inventory Analysis sheet (no data available)...")
            placeholder_df = pd.DataFrame([
                ['Inventory Analysis', 'No inventory data available'],
                ['Status', 'Please ensure InventoryData sheet is present in uploaded file']
            ], columns=['Item', 'Value'])
            placeholder_df.to_excel(writer, sheet_name='Inventory_Analysis', index=False)
            return
        
        if self.verbose:
            print("Creating Inventory Analysis sheet...")
        
        inventory_results = self.analysis_results['inventory_analysis']
        
        # Get the SKU inventory matrix and daily summary
        sku_matrix = inventory_results.get('sku_inventory_matrix')
        daily_summary = inventory_results.get('daily_summary')
        
        if sku_matrix is not None and not sku_matrix.empty:
            # Write SKU inventory matrix (main table on left)
            sku_matrix.to_excel(writer, sheet_name='Inventory_Analysis', 
                              startrow=0, startcol=0, index=False)
            
            # Calculate position for daily summary table (to the right)
            # Add 2 columns of spacing after the SKU matrix
            right_col_start = len(sku_matrix.columns) + 2
            
            if daily_summary is not None and not daily_summary.empty:
                # Write daily summary table to the right
                daily_summary.to_excel(writer, sheet_name='Inventory_Analysis',
                                     startrow=0, startcol=right_col_start, index=False)

                # Write Inventory Health Metrics (aggregate KPIs) to the right of daily summary
                inventory_statistics = inventory_results.get('inventory_statistics', {})
                health = inventory_statistics.get('inventory_health', {})
                if health:
                    health_start_col = right_col_start + 4
                    health_data = [
                        [k.replace('_', ' ').title(), str(v)]
                        for k, v in health.items()
                    ]
                    health_df = pd.DataFrame(health_data, columns=['Metric', 'Value'])
                    health_df.to_excel(writer, sheet_name='Inventory_Analysis',
                                       startrow=0, startcol=health_start_col, index=False)

                    # ── Charts ───────────────────────────────────────────────
                    try:
                        from charts.excel_chart_generator import ExcelChartGenerator
                        ws_inv    = writer.sheets['Inventory_Analysis']
                        chart_gen = ExcelChartGenerator(ws_inv)

                        # Chart 1: Daily stock trend line chart
                        num_daily_rows = len(daily_summary)
                        if num_daily_rows > 1:
                            daily_col_excel = right_col_start + 1  # pandas 0-based → 1-based
                            # Anchor right of health metrics (2 data cols + 2 gap, 1-based)
                            chart1_col = get_column_letter(health_start_col + 2 + 2 + 1)
                            chart_gen.add_inventory_stock_trend_chart(
                                daily_start_col=daily_col_excel,
                                num_daily_rows=num_daily_rows,
                                anchor=f"{chart1_col}1"
                            )

                        # Chart 2: Stock status pie chart
                        # Write mini 2-column helper table below health_df, then chart
                        pie_data = pd.DataFrame({
                            'Status':    ['Low Stock', 'Excess Stock', 'OK', 'No Demand'],
                            'SKU Count': [
                                health.get('skus_low_stock', 0),
                                health.get('skus_excess_stock', 0),
                                health.get('skus_ok', 0),
                                health.get('skus_no_demand', 0)
                            ]
                        })
                        pie_row_pandas      = len(health_df) + 2    # below health_df + 1 blank row
                        pie_data.to_excel(writer, sheet_name='Inventory_Analysis',
                                          startrow=pie_row_pandas, startcol=health_start_col,
                                          index=False)
                        pie_start_col_excel = health_start_col + 1  # 1-based
                        pie_start_row_excel = pie_row_pandas + 1     # 1-based
                        chart2_col = get_column_letter(health_start_col + 2 + 2 + 1)
                        chart_gen.add_inventory_stock_status_chart(
                            pie_start_col=pie_start_col_excel,
                            pie_start_row=pie_start_row_excel,
                            num_rows=len(pie_data),
                            anchor=f"{chart2_col}13"   # below stock trend chart (~12 rows)
                        )
                    except Exception as e:
                        print(f"Warning: Inventory chart generation failed: {e}")

                if self.verbose:
                    print(f"  Created inventory analysis with {len(sku_matrix)} SKUs")
                    print(f"  Daily summary placed at column {right_col_start}")
        else:
            # No inventory data to display
            if self.verbose:
                print("  No inventory matrix data available")
            placeholder_df = pd.DataFrame([
                ['Inventory Analysis', 'No data to display'],
                ['Note', 'Inventory data may be missing or filtered out']
            ], columns=['Item', 'Value'])
            placeholder_df.to_excel(writer, sheet_name='Inventory_Analysis', index=False)
    
    def _create_manpower_analysis_sheet(self, writer):
        """Create manpower analysis sheet with placeholder data"""
        if 'manpower_analysis' not in self.analysis_results:
            # Create placeholder if no manpower analysis results
            if self.verbose:
                print("Creating Manpower Analysis sheet (no data available)...")
            placeholder_df = pd.DataFrame([
                ['Manpower Analysis', 'No analysis performed'],
                ['Status', 'Order or receipt data required for manpower analysis'],
                ['Implementation', 'Placeholder - full implementation pending']
            ], columns=['Item', 'Value'])
            placeholder_df.to_excel(writer, sheet_name='Manpower_Analysis', index=False)
            return
        
        if self.verbose:
            print("Creating Manpower Analysis sheet...")
        
        manpower_results = self.analysis_results['manpower_analysis']
        current_row = 0
        # Chart position dicts — set when each table is written, used at the end
        table_pos_pt          = None
        table_pos_hourly_pick = None
        table_pos_hourly_recv = None

        # Section 1: Picking Analysis
        picking_analysis = manpower_results.get('picking_analysis', {})
        if picking_analysis:
            # Add main section header
            header_df = pd.DataFrame([['PICKING MANPOWER ANALYSIS']], columns=['Section'])
            header_df.to_excel(writer, sheet_name='Manpower_Analysis', startrow=current_row, index=False, header=False)
            current_row += 2
            
            # Subsection 1A: Daily Summary
            daily_summary = picking_analysis.get('daily_summary', {})
            if daily_summary:
                # Daily summary header
                summary_header = pd.DataFrame([['Daily Staffing Summary']], columns=['Section'])
                summary_header.to_excel(writer, sheet_name='Manpower_Analysis', startrow=current_row, index=False, header=False)
                current_row += 1
                
                # Daily summary data
                daily_data = []
                for key, value in daily_summary.items():
                    formatted_key = key.replace('_', ' ').title()
                    daily_data.append([formatted_key, str(value)])
                
                daily_df = pd.DataFrame(daily_data, columns=['Metric', 'Value'])
                daily_df.to_excel(writer, sheet_name='Manpower_Analysis', startrow=current_row, index=False)
                current_row += len(daily_df) + 2
            
            # Subsection 1D: Pick Type Breakdown
            pick_type_breakdown = picking_analysis.get('pick_type_breakdown', {})
            if pick_type_breakdown:
                pt_header = pd.DataFrame([['Pick Type Breakdown']], columns=['Section'])
                pt_header.to_excel(writer, sheet_name='Manpower_Analysis', startrow=current_row, index=False, header=False)
                current_row += 1

                pt_df = pd.DataFrame.from_dict(pick_type_breakdown, orient='index').reset_index()
                pt_df.columns = ['Pick Type', 'Order Lines', 'Total Cases', 'Total Eaches']
                pt_start_row = current_row
                pt_df.to_excel(writer, sheet_name='Manpower_Analysis', startrow=current_row, index=False)
                current_row += len(pt_df) + 2
                table_pos_pt = {
                    'row': pt_start_row + 1,   # 1-based Excel row of header
                    'col': 1,
                    'num_rows': len(pt_df),
                    'num_cols': len(pt_df.columns)
                }

            # Subsection 1B: Shift Breakdown
            shift_breakdown = picking_analysis.get('shift_breakdown', [])
            if shift_breakdown:
                # Shift breakdown header
                shift_header = pd.DataFrame([['Shift-wise Breakdown']], columns=['Section'])
                shift_header.to_excel(writer, sheet_name='Manpower_Analysis', startrow=current_row, index=False, header=False)
                current_row += 1
                
                # Convert shift breakdown to DataFrame
                shift_df = pd.DataFrame(shift_breakdown)
                if not shift_df.empty:
                    shift_df.to_excel(writer, sheet_name='Manpower_Analysis', startrow=current_row, index=False)
                    current_row += len(shift_df) + 2
            
            # Subsection 1C: Hourly Requirements
            hourly_requirements = picking_analysis.get('hourly_requirements', [])
            if hourly_requirements:
                # Hourly requirements header
                hourly_header = pd.DataFrame([['Hourly Requirements']], columns=['Section'])
                hourly_header.to_excel(writer, sheet_name='Manpower_Analysis', startrow=current_row, index=False, header=False)
                current_row += 1
                
                # Convert hourly requirements to DataFrame
                hourly_df = pd.DataFrame(hourly_requirements)
                if not hourly_df.empty:
                    hourly_pick_start_row = current_row
                    hourly_df.to_excel(writer, sheet_name='Manpower_Analysis', startrow=current_row, index=False)
                    current_row += len(hourly_df) + 3
                    table_pos_hourly_pick = {
                        'row': hourly_pick_start_row + 1,
                        'col': 1,
                        'num_rows': len(hourly_df),
                        'num_cols': len(hourly_df.columns)
                    }

        # Section 2: Receiving Analysis
        receiving_analysis = manpower_results.get('receiving_analysis', {})
        if receiving_analysis:
            # Add main section header
            header_df = pd.DataFrame([['RECEIVING & PUTAWAY MANPOWER ANALYSIS']], columns=['Section'])
            header_df.to_excel(writer, sheet_name='Manpower_Analysis', startrow=current_row, index=False, header=False)
            current_row += 2
            
            # Check if this is simplified analysis with three subsections
            daily_summary = receiving_analysis.get('daily_summary', {})
            if daily_summary:
                # Subsection 2A: Daily Summary
                summary_header = pd.DataFrame([['Daily Staffing Summary']], columns=['Section'])
                summary_header.to_excel(writer, sheet_name='Manpower_Analysis', startrow=current_row, index=False, header=False)
                current_row += 1
                
                # Daily summary data
                daily_data = []
                for key, value in daily_summary.items():
                    formatted_key = key.replace('_', ' ').title()
                    daily_data.append([formatted_key, str(value)])
                
                daily_df = pd.DataFrame(daily_data, columns=['Metric', 'Value'])
                daily_df.to_excel(writer, sheet_name='Manpower_Analysis', startrow=current_row, index=False)
                current_row += len(daily_df) + 2
                
                # Subsection 2D: Truck Analysis
                truck_analysis = receiving_analysis.get('truck_analysis', {})
                if truck_analysis:
                    ta_header = pd.DataFrame([['Truck Analysis']], columns=['Section'])
                    ta_header.to_excel(writer, sheet_name='Manpower_Analysis', startrow=current_row, index=False, header=False)
                    current_row += 1

                    ta_data = [[k.replace('_', ' ').title(), str(v)] for k, v in truck_analysis.items()]
                    ta_df = pd.DataFrame(ta_data, columns=['Metric', 'Value'])
                    ta_df.to_excel(writer, sheet_name='Manpower_Analysis', startrow=current_row, index=False)
                    current_row += len(ta_df) + 2

                # Subsection 2B: Shift Breakdown
                shift_breakdown = receiving_analysis.get('shift_breakdown', [])
                if shift_breakdown:
                    shift_header = pd.DataFrame([['Shift-wise Breakdown']], columns=['Section'])
                    shift_header.to_excel(writer, sheet_name='Manpower_Analysis', startrow=current_row, index=False, header=False)
                    current_row += 1
                    
                    shift_df = pd.DataFrame(shift_breakdown)
                    if not shift_df.empty:
                        shift_df.to_excel(writer, sheet_name='Manpower_Analysis', startrow=current_row, index=False)
                        current_row += len(shift_df) + 2
                
                # Subsection 2C: Hourly Requirements
                hourly_requirements = receiving_analysis.get('hourly_requirements', [])
                if hourly_requirements:
                    hourly_header = pd.DataFrame([['Hourly Requirements']], columns=['Section'])
                    hourly_header.to_excel(writer, sheet_name='Manpower_Analysis', startrow=current_row, index=False, header=False)
                    current_row += 1
                    
                    hourly_df = pd.DataFrame(hourly_requirements)
                    if not hourly_df.empty:
                        hourly_recv_start_row = current_row
                        hourly_df.to_excel(writer, sheet_name='Manpower_Analysis', startrow=current_row, index=False)
                        current_row += len(hourly_df) + 3
                        table_pos_hourly_recv = {
                            'row': hourly_recv_start_row + 1,
                            'col': 1,
                            'num_rows': len(hourly_df),
                            'num_cols': len(hourly_df.columns)
                        }
            else:
                # Legacy format - simple key-value pairs
                receiving_data = []
                for key, value in receiving_analysis.items():
                    if key != 'notes':
                        receiving_data.append([key.replace('_', ' ').title(), str(value)])
                
                if receiving_data:
                    receiving_df = pd.DataFrame(receiving_data, columns=['Metric', 'Value'])
                    receiving_df.to_excel(writer, sheet_name='Manpower_Analysis', startrow=current_row, index=False)
                    current_row += len(receiving_df) + 3
        
        # Section 3: Loading Analysis
        loading_analysis = manpower_results.get('loading_analysis', {})
        if loading_analysis:
            # Add section header
            header_df = pd.DataFrame([['LOADING MANPOWER ANALYSIS']], columns=['Section'])
            header_df.to_excel(writer, sheet_name='Manpower_Analysis', startrow=current_row, index=False, header=False)
            current_row += 2
            
            # Create loading analysis data
            loading_data = []
            for key, value in loading_analysis.items():
                if key != 'notes':
                    loading_data.append([key.replace('_', ' ').title(), str(value)])
            
            if loading_data:
                loading_df = pd.DataFrame(loading_data, columns=['Metric', 'Value'])
                loading_df.to_excel(writer, sheet_name='Manpower_Analysis', startrow=current_row, index=False)
                current_row += len(loading_df) + 3
        
        # Section 4: Efficiency Summary
        efficiency_summary = manpower_results.get('efficiency_summary', {})
        if efficiency_summary:
            # Add section header
            header_df = pd.DataFrame([['EFFICIENCY SUMMARY & RECOMMENDATIONS']], columns=['Section'])
            header_df.to_excel(writer, sheet_name='Manpower_Analysis', startrow=current_row, index=False, header=False)
            current_row += 2
            
            # Create efficiency summary data
            efficiency_data = []
            for key, value in efficiency_summary.items():
                if key not in ['optimization_opportunities', 'cost_analysis', 'notes']:
                    efficiency_data.append([key.replace('_', ' ').title(), str(value)])
            
            if efficiency_data:
                efficiency_df = pd.DataFrame(efficiency_data, columns=['Metric', 'Value'])
                efficiency_df.to_excel(writer, sheet_name='Manpower_Analysis', startrow=current_row, index=False)
                current_row += len(efficiency_df) + 2
            
            # Add optimization opportunities
            if 'optimization_opportunities' in efficiency_summary:
                header_df = pd.DataFrame([['OPTIMIZATION OPPORTUNITIES']], columns=['Section'])
                header_df.to_excel(writer, sheet_name='Manpower_Analysis', startrow=current_row, index=False, header=False)
                current_row += 2
                
                opportunities = efficiency_summary['optimization_opportunities']
                if opportunities:
                    opp_data = [[f"{i+1}. {opp}"] for i, opp in enumerate(opportunities)]
                    opp_df = pd.DataFrame(opp_data, columns=['Recommendations'])
                    opp_df.to_excel(writer, sheet_name='Manpower_Analysis', startrow=current_row, index=False)

        # ── Charts ───────────────────────────────────────────────────────────
        try:
            from charts.excel_chart_generator import ExcelChartGenerator
            ws_mp      = writer.sheets['Manpower_Analysis']
            chart_gen_mp = ExcelChartGenerator(ws_mp)

            if table_pos_pt is not None:
                chart_gen_mp.add_manpower_pick_type_chart(table_pos_pt)

            if table_pos_hourly_pick is not None:
                chart_gen_mp.add_manpower_hourly_profile_chart(
                    table_pos_hourly_pick, section_title='Picking')

            if table_pos_hourly_recv is not None:
                chart_gen_mp.add_manpower_hourly_profile_chart(
                    table_pos_hourly_recv, section_title='Receiving')
        except Exception as e:
            print(f"Warning: Manpower chart generation failed: {e}")

    def _create_receipt_analysis_sheet(self, writer):
        """Create receipt analysis sheet - daily patterns with percentile analysis on the right"""
        if 'receipt_analysis' not in self.analysis_results:
            return
        
        if self.verbose:
            print("Creating Receipt Analysis sheet...")
        
        receipt_results = self.analysis_results['receipt_analysis']
        
        # Daily patterns on the left
        daily_patterns = receipt_results.get('daily_patterns', {})
        if 'daily_data' in daily_patterns:
            daily_df = daily_patterns['daily_data']
            daily_df.to_excel(writer, sheet_name='Receipt_Analysis', startrow=0, startcol=0, index=False)
            
            # Add receipt charts
            from charts.excel_chart_generator import ExcelChartGenerator
            chart_gen = ExcelChartGenerator(writer.sheets['Receipt_Analysis'])
            
            table_pos = {
                'row': 1,  # Excel uses 1-based indexing
                'col': 1,
                'num_rows': len(daily_df),
                'num_cols': len(daily_df.columns)
            }
            
            # Add main receipt trend chart (Lines, Shipments, Trucks) - place further right to avoid percentile overlap
            chart_gen.add_receipt_daily_trend_chart(table_pos, columns_gap=12)
            
            # Add volume trend chart below the first chart
            chart_gen.add_receipt_volume_trend_chart(table_pos, columns_gap=12)
            
            # Calculate the right column position (after daily patterns table + some spacing)
            right_col_start = len(daily_df.columns) + 2
            
            # Percentile Analysis on the right
            percentile_analysis = receipt_results.get('percentile_analysis', {})
            if percentile_analysis:
                current_row = 0

                # Header for percentile analysis
                perc_header = pd.DataFrame([['RECEIPT PERCENTILE ANALYSIS']], columns=['Header'])
                perc_header.to_excel(writer, sheet_name='Receipt_Analysis', startrow=current_row, startcol=right_col_start, index=False, header=False)
                current_row += 2

                # Column headers for percentile table (7 columns — added Case Equiv Vol)
                perc_col_names = ['Percentile', '#Trucks', '#Shipments', '#Lines', '#SKUs', '#Cases', 'Case Equiv Vol']
                headers = pd.DataFrame([perc_col_names], columns=perc_col_names)
                headers.to_excel(writer, sheet_name='Receipt_Analysis', startrow=current_row, startcol=right_col_start, index=False, header=False)
                current_row += 1

                # Get horizontal percentile data
                horizontal_percentiles = percentile_analysis.get('horizontal_percentiles', {})

                if horizontal_percentiles:
                    percentile_data = []

                    # Build row order from available percentiles
                    row_order = []
                    if 'Max' in horizontal_percentiles:
                        row_order.append('Max')

                    # Add all configured percentiles in descending order
                    percentile_keys = [k for k in horizontal_percentiles.keys() if k.endswith('%ile')]
                    percentile_keys.sort(key=lambda x: float(x.replace('%ile', '')), reverse=True)
                    row_order.extend(percentile_keys)

                    for percentile_name in row_order:
                        if percentile_name in horizontal_percentiles:
                            row_data = horizontal_percentiles[percentile_name]
                            percentile_data.append([
                                percentile_name,
                                int(row_data.get('#Trucks', 0)),
                                int(row_data.get('#Shipments', 0)),
                                int(row_data.get('#Lines', 0)),
                                int(row_data.get('#SKUs', 0)),
                                int(row_data.get('#Cases', 0)),
                                round(row_data.get('Case_Equiv_Volume', 0), 0)
                            ])

                    perc_df = pd.DataFrame(percentile_data, columns=perc_col_names)
                    perc_df.to_excel(writer, sheet_name='Receipt_Analysis', startrow=current_row, startcol=right_col_start, index=False, header=False)
                    current_row += len(perc_df) + 3

                # ── C2: Truck / Supplier Performance Summary ──────────────────────────
                supplier_performance = receipt_results.get('supplier_performance', {})
                perf_summary = supplier_performance.get('performance_summary')
                if perf_summary is not None and not perf_summary.empty:
                    perf_h = pd.DataFrame([['TRUCK PERFORMANCE SUMMARY']])
                    perf_h.to_excel(writer, sheet_name='Receipt_Analysis',
                                    startrow=current_row, startcol=right_col_start,
                                    index=False, header=False)
                    current_row += 1
                    perf_summary.reset_index().to_excel(
                        writer, sheet_name='Receipt_Analysis',
                        startrow=current_row, startcol=right_col_start, index=False)
                    current_row += len(perf_summary) + 3

                # ── C3: Dock Utilization Summary ──────────────────────────────────────
                dock_utilization = receipt_results.get('dock_utilization', {})
                util_summary = dock_utilization.get('utilization_summary')
                if util_summary is not None and not util_summary.empty:
                    # KPI row first
                    kpi_data = [
                        ['Avg Dock Utilization %', round(dock_utilization.get('avg_utilization', 0), 1)],
                        ['Over-Capacity Days', dock_utilization.get('over_capacity_days', 0)],
                        ['Max Trucks Assumption', dock_utilization.get('max_capacity_assumption', 0)],
                    ]
                    kpi_df = pd.DataFrame(kpi_data, columns=['Metric', 'Value'])
                    dock_h = pd.DataFrame([['DOCK UTILIZATION SUMMARY']])
                    dock_h.to_excel(writer, sheet_name='Receipt_Analysis',
                                    startrow=current_row, startcol=right_col_start,
                                    index=False, header=False)
                    current_row += 1
                    kpi_df.to_excel(writer, sheet_name='Receipt_Analysis',
                                    startrow=current_row, startcol=right_col_start, index=False)
                    current_row += len(kpi_df) + 1
                    util_summary.reset_index().to_excel(
                        writer, sheet_name='Receipt_Analysis',
                        startrow=current_row, startcol=right_col_start, index=False)
                    current_row += len(util_summary) + 3

                # ── C4: Receiving Efficiency ───────────────────────────────────────────
                receiving_efficiency = receipt_results.get('receiving_efficiency', {})
                efficiency_stats = receiving_efficiency.get('efficiency_stats', {})
                if efficiency_stats:
                    eff_h = pd.DataFrame([['RECEIVING EFFICIENCY']])
                    eff_h.to_excel(writer, sheet_name='Receipt_Analysis',
                                   startrow=current_row, startcol=right_col_start,
                                   index=False, header=False)
                    current_row += 1
                    eff_data = [
                        [k.replace('_', ' ').title(),
                         round(v, 2) if isinstance(v, float) else v]
                        for k, v in efficiency_stats.items()
                    ]
                    eff_df = pd.DataFrame(eff_data, columns=['Metric', 'Value'])
                    eff_df.to_excel(writer, sheet_name='Receipt_Analysis',
                                    startrow=current_row, startcol=right_col_start, index=False)
                    current_row += len(eff_df) + 3

                # ── C5: Inter-Receipt Interval (Replenishment Lead Time Proxy) ─────────
                lead_times = receipt_results.get('lead_times', {})
                if 'avg_inter_receipt_days' in lead_times:
                    lt_h = pd.DataFrame([['INTER-RECEIPT INTERVAL (REPLENISHMENT LEAD TIME PROXY)']])
                    lt_h.to_excel(writer, sheet_name='Receipt_Analysis',
                                  startrow=current_row, startcol=right_col_start,
                                  index=False, header=False)
                    current_row += 1
                    lt_data = [
                        ['Avg Inter-Receipt Days',    lead_times['avg_inter_receipt_days']],
                        ['Median Inter-Receipt Days', lead_times['median_inter_receipt_days']],
                        ['P95 Inter-Receipt Days',    lead_times['p95_inter_receipt_days']],
                        ['Std Inter-Receipt Days',    lead_times['std_inter_receipt_days']],
                    ]
                    lt_df = pd.DataFrame(lt_data, columns=['Metric', 'Value'])
                    lt_df.to_excel(writer, sheet_name='Receipt_Analysis',
                                   startrow=current_row, startcol=right_col_start, index=False)
                    current_row += len(lt_df) + 3

                # ── C6: SKU Receipt Pattern Summary ───────────────────────────────────
                sku_patterns = receipt_results.get('sku_patterns', {})
                pattern_summary = sku_patterns.get('pattern_summary')
                if pattern_summary is not None and not pattern_summary.empty:
                    pat_h = pd.DataFrame([['SKU RECEIPT PATTERN SUMMARY']])
                    pat_h.to_excel(writer, sheet_name='Receipt_Analysis',
                                   startrow=current_row, startcol=right_col_start,
                                   index=False, header=False)
                    current_row += 1
                    pattern_summary.reset_index().to_excel(
                        writer, sheet_name='Receipt_Analysis',
                        startrow=current_row, startcol=right_col_start, index=False)
    
    def _create_recommendations_sheet(self, writer):
        """Create consolidated recommendations sheet"""
        if self.verbose:
            print("Creating Recommendations sheet...")
        
        all_recommendations = []
        
        # Collect recommendations from all analyses
        analysis_modules = ['order_analysis', 'sku_analysis', 'abc_fms_analysis', 'receipt_analysis']
        
        for module in analysis_modules:
            if module in self.analysis_results:
                module_results = self.analysis_results[module]
                
                # Check for recommendations in different possible locations
                recommendations = None
                if 'recommendations' in module_results:
                    recs_data = module_results['recommendations']
                    if isinstance(recs_data, dict) and 'recommendations' in recs_data:
                        recommendations = recs_data['recommendations']
                    elif isinstance(recs_data, list):
                        recommendations = recs_data
                
                # Also check for strategic recommendations
                if 'strategic_recommendations' in module_results:
                    strategic_recs = module_results['strategic_recommendations']
                    if isinstance(strategic_recs, dict) and 'recommendations' in strategic_recs:
                        strategic_recommendations = strategic_recs['recommendations']
                        if recommendations:
                            recommendations.extend(strategic_recommendations)
                        else:
                            recommendations = strategic_recommendations
                
                if recommendations:
                    for rec in recommendations:
                        all_recommendations.append([
                            module.replace('_', ' ').title(),
                            rec.get('category', rec.get('segment', 'General')),
                            rec.get('priority', 'Medium'),
                            rec.get('recommendation', ''),
                            rec.get('impact', rec.get('actions', ''))
                        ])
        
        if all_recommendations:
            recs_df = pd.DataFrame(all_recommendations, 
                                 columns=['Analysis Module', 'Category', 'Priority', 'Recommendation', 'Expected Impact'])
            
            # Sort by priority
            priority_order = {'Critical': 1, 'High': 2, 'Medium': 3, 'Low': 4}
            recs_df['Priority_Order'] = recs_df['Priority'].map(priority_order).fillna(5)
            recs_df = recs_df.sort_values('Priority_Order').drop('Priority_Order', axis=1)
            
            recs_df.to_excel(writer, sheet_name='Recommendations', index=False)
        else:
            # Create placeholder if no recommendations found
            placeholder_df = pd.DataFrame([['No recommendations available', '', '', '', '']], 
                                        columns=['Analysis Module', 'Category', 'Priority', 'Recommendation', 'Expected Impact'])
            placeholder_df.to_excel(writer, sheet_name='Recommendations', index=False)
    
    def _create_configuration_sheet(self, writer):
        """Create configuration documentation sheet"""
        if self.verbose:
            print("Creating Configuration sheet...")
        
        config_data = []
        
        # Analysis configuration
        config_data.append(['ANALYSIS CONFIGURATION', ''])
        config_data.append(['Configuration Generated', datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
        config_data.append(['', ''])
        
        if self.configuration:
            for category, settings in self.configuration.items():
                config_data.append([f'{category.upper().replace("_", " ")}', ''])
                if isinstance(settings, dict):
                    for key, value in settings.items():
                        config_data.append([key.replace('_', ' ').title(), str(value)])
                else:
                    config_data.append(['Value', str(settings)])
                config_data.append(['', ''])
        
        # Add default settings from config module
        config_data.append(['DEFAULT SETTINGS', ''])
        config_data.append(['ABC A Threshold', f"{config.DEFAULT_ABC_THRESHOLDS['A_THRESHOLD']}%"])
        config_data.append(['ABC B Threshold', f"{config.DEFAULT_ABC_THRESHOLDS['B_THRESHOLD']}%"])
        config_data.append(['FMS Fast Threshold', f"{config.DEFAULT_FMS_THRESHOLDS['F_THRESHOLD']}%"])
        config_data.append(['FMS Medium Threshold', f"{config.DEFAULT_FMS_THRESHOLDS['M_THRESHOLD']}%"])
        config_data.append(['Default Percentiles', ', '.join(map(str, config.DEFAULT_PERCENTILE_LEVELS))])
        
        config_df = pd.DataFrame(config_data, columns=['Setting', 'Value'])
        config_df.to_excel(writer, sheet_name='Configuration', index=False)
    
    def _create_raw_data_summary(self, writer):
        """Create raw data summary sheet"""
        if self.verbose:
            print("Creating Raw Data Summary sheet...")
        
        summary_data = []
        
        # Data source information
        summary_data.append(['RAW DATA SUMMARY', ''])
        summary_data.append(['Report Generated', datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
        summary_data.append(['Tool Version', 'Warehouse Analysis Tool V2'])
        summary_data.append(['', ''])
        
        # Data availability summary
        if 'data_loader' in self.analysis_results:
            data_results = self.analysis_results['data_loader']
            
            summary_data.append(['DATA AVAILABILITY', ''])
            
            for data_type, data_info in data_results.get('data', {}).items():
                if isinstance(data_info, pd.DataFrame):
                    summary_data.append([data_type.replace('_', ' ').title(), f'{len(data_info)} records'])
                else:
                    summary_data.append([data_type.replace('_', ' ').title(), 'Available'])
            
            summary_data.append(['', ''])
            
            # Validation results
            validation_results = data_results.get('validation', {})
            if validation_results:
                summary_data.append(['DATA VALIDATION', ''])
                for data_type, validation in validation_results.items():
                    if isinstance(validation, dict):
                        summary_data.append([f'{data_type.replace("_", " ").title()} - Total Rows', validation.get('total_rows', 'N/A')])
                        summary_data.append([f'{data_type.replace("_", " ").title()} - Date Range (Days)', validation.get('date_range_days', 'N/A')])
        
        # Analysis modules executed
        summary_data.append(['', ''])
        summary_data.append(['ANALYSIS MODULES EXECUTED', ''])
        
        executed_modules = [key for key in self.analysis_results.keys() if key != 'data_loader']
        for module in executed_modules:
            summary_data.append([module.replace('_', ' ').title(), 'Completed'])
        
        summary_df = pd.DataFrame(summary_data, columns=['Item', 'Value'])
        summary_df.to_excel(writer, sheet_name='Raw_Data_Summary', index=False)
    
# Test function for standalone execution
if __name__ == "__main__":
    print("ExcelGenerator module - ready for use")
    print("This module requires analysis results to function.")
    print("Use within the main analysis pipeline for proper functionality.")