#!/usr/bin/env python3
"""
Excel Chart Generator - Central coordinator for chart placement in Excel workbooks

This module handles the placement of charts in Excel workbooks, managing
the creation of References and positioning of charts based on table locations.
"""

from openpyxl.chart import Reference, LineChart, BarChart, PieChart
from openpyxl.chart.series import DataPoint
from openpyxl.utils import get_column_letter
from typing import Dict, Optional
import pandas as pd

# Import all chart modules
from .order_analysis_charts import OrderAnalysisCharts
from .receipt_analysis_charts import ReceiptAnalysisCharts
from .abc_fms_analysis_charts import ABCFMSAnalysisCharts
# from .sku_analysis_charts import SKUAnalysisCharts
# from .inventory_analysis_charts import InventoryAnalysisCharts
# from .manpower_analysis_charts import ManpowerAnalysisCharts


class ExcelChartGenerator:
    """
    Coordinates chart placement in Excel workbooks.
    Handles Reference creation and positioning logic.
    """
    
    def __init__(self, worksheet):
        """
        Initialize with a worksheet object.
        
        Args:
            worksheet: openpyxl worksheet object
        """
        self.ws = worksheet
        
        # Initialize all chart modules
        self.order_charts = OrderAnalysisCharts()
        self.receipt_charts = ReceiptAnalysisCharts()
        self.abc_fms_charts = ABCFMSAnalysisCharts()
        
    def add_order_daily_trend_chart(self, table_position: Dict, columns_gap: int = 2) -> bool:
        """
        Add Order Profile daily trend chart to the right of the daily data table.
        
        Args:
            table_position: Dictionary with table location info:
                - 'row': Starting row of table (1-based Excel indexing)
                - 'col': Starting column of table (1-based)
                - 'num_rows': Number of data rows (excluding header)
                - 'num_cols': Number of columns in table
            columns_gap: Number of columns to leave between table and chart (default 2)
            
        Returns:
            bool: True if chart was added successfully, False otherwise
        """
        try:
            # Extract position info
            start_row = table_position['row']
            start_col = table_position['col']
            num_rows = table_position['num_rows']
            
            # Calculate rows
            header_row = start_row
            first_data_row = start_row + 1
            last_data_row = start_row + num_rows
            
            # Create References for the data
            # Map to actual columns: Date, Daily_Order_Lines, Daily_Orders, Daily_Shipments, Daily_SKUs
            
            # Dates (column A, no header)
            dates_ref = Reference(self.ws, 
                                 min_col=start_col, 
                                 min_row=first_data_row, 
                                 max_row=last_data_row)
            
            # Daily_Orders (column C, with header for series name)
            lines_ref = Reference(self.ws, 
                                min_col=start_col + 2,
                                min_row=header_row, 
                                max_row=last_data_row)
            
            # Daily_Shipments (column D, with header)
            customers_ref = Reference(self.ws, 
                                    min_col=start_col + 3,
                                    min_row=header_row, 
                                    max_row=last_data_row)
            
            # Daily_SKUs (column E, with header)
            shipments_ref = Reference(self.ws, 
                                    min_col=start_col + 4,
                                    min_row=header_row, 
                                    max_row=last_data_row)
            
            # Get chart from chart module
            chart = self.order_charts.create_daily_trend_chart(
                ws=self.ws,
                dates_ref=dates_ref,
                lines_ref=lines_ref,
                customers_ref=customers_ref,
                shipments_ref=shipments_ref
            )
            
            # Calculate chart position (to the right of table)
            chart_col = start_col + table_position['num_cols'] + columns_gap
            chart_position = f"{get_column_letter(chart_col)}{start_row}"
            
            # Add chart to worksheet
            self.ws.add_chart(chart, chart_position)
            
            print(f"✅ Order trend chart added at {chart_position}")
            return True
            
        except Exception as e:
            print(f"⚠️ Could not add order trend chart: {str(e)}")
            return False
    
    def add_order_volume_trend_chart(self, table_position: Dict, columns_gap: int = 2, 
                                   row_gap: int = 22) -> bool:
        """
        Add Daily Case Equivalent Volume trend chart below the main trend chart.
        
        Args:
            table_position: Dictionary with table location info
            columns_gap: Number of columns to leave between table and chart (default 2)
            row_gap: Number of rows below the first chart (default 12)
            
        Returns:
            bool: True if chart was added successfully, False otherwise
        """
        try:
            # Extract position info
            start_row = table_position['row']
            start_col = table_position['col']
            num_rows = table_position['num_rows']
            
            # Calculate rows
            header_row = start_row
            first_data_row = start_row + 1
            last_data_row = start_row + num_rows
            
            # Create References for the data
            # Dates (column A, no header)
            dates_ref = Reference(self.ws, 
                                 min_col=start_col, 
                                 min_row=first_data_row, 
                                 max_row=last_data_row)
            
            # Daily_Order_Lines (column B, with header)
            lines_ref = Reference(self.ws, 
                                min_col=start_col + 1,  # Column B
                                min_row=header_row, 
                                max_row=last_data_row)
            
            # Daily_Case_Equivalent_Volume (column H, with header)
            volume_ref = Reference(self.ws, 
                                  min_col=start_col + 7,  # Column H (8th column)
                                  min_row=header_row, 
                                  max_row=last_data_row)
            
            # Get chart from chart module
            chart = self.order_charts.create_volume_trend_chart(
                ws=self.ws,
                dates_ref=dates_ref,
                lines_ref=lines_ref,
                volume_ref=volume_ref
            )
            
            # Calculate chart position (below the first chart)
            chart_col = start_col + table_position['num_cols'] + columns_gap
            chart_row = start_row + row_gap  # Position below first chart
            chart_position = f"{get_column_letter(chart_col)}{chart_row}"
            
            # Add chart to worksheet
            self.ws.add_chart(chart, chart_position)
            
            print(f"✅ Order volume chart added at {chart_position}")
            return True
            
        except Exception as e:
            print(f"⚠️ Could not add order volume chart: {str(e)}")
            return False
    
    def add_order_percentile_chart(self, table_position: Dict, columns_gap: int = 2) -> bool:
        """
        Add percentile analysis column chart (Total_Case_Equiv per percentile level).

        Args:
            table_position: Dictionary with table location info:
                - 'row': Header row of percentile table (1-based Excel indexing)
                - 'col': Starting column (1-based)
                - 'num_rows': Number of data rows (excluding header)
                - 'num_cols': Number of columns in table
            columns_gap: Gap between table and chart

        Returns:
            bool: Success status
        """
        try:
            start_row  = table_position['row']
            num_rows   = table_position['num_rows']
            header_row = start_row
            first_data = start_row + 1
            last_data  = start_row + num_rows

            # Percentile labels (col 1, data rows only — no header needed as categories)
            categories_ref = Reference(self.ws, min_col=1,
                                       min_row=first_data, max_row=last_data)
            # Total_Case_Equiv is column 8; include header row for series name
            case_equiv_ref = Reference(self.ws, min_col=8,
                                       min_row=header_row, max_row=last_data)

            chart = self.order_charts.create_percentile_chart(
                self.ws, categories_ref, [case_equiv_ref]
            )
            if chart:
                chart_position = self._calculate_chart_position(
                    table_position, placement='right', gap=columns_gap
                )
                self.ws.add_chart(chart, chart_position)

            print(f"✅ Order percentile chart added")
            return True

        except Exception as e:
            print(f"⚠️ Could not add order percentile chart: {str(e)}")
            return False
    
    def add_receipt_daily_trend_chart(self, table_position: Dict, columns_gap: int = 2) -> bool:
        """
        Add Receipt Profile daily trend chart.
        
        Args:
            table_position: Dictionary with table location info
            columns_gap: Gap between table and chart
            
        Returns:
            bool: Success status
        """
        try:
            # Extract position info
            start_row = table_position['row']
            start_col = table_position['col']
            num_rows = table_position['num_rows']
            
            # Calculate rows
            header_row = start_row
            first_data_row = start_row + 1
            last_data_row = start_row + num_rows
            
            # Create References for the data
            # Map to actual columns: Date, Daily_Receipt_Lines, Daily_SKUs, Daily_Shipments, Daily_Trucks
            
            # Dates (column A, no header)
            dates_ref = Reference(self.ws, 
                                 min_col=start_col, 
                                 min_row=first_data_row, 
                                 max_row=last_data_row)
            
            # Daily_Receipt_Lines (column B, with header for series name)
            lines_ref = Reference(self.ws, 
                                min_col=start_col + 1,
                                min_row=header_row, 
                                max_row=last_data_row)
            
            # Daily_Shipments (column D, with header)
            shipments_ref = Reference(self.ws, 
                                    min_col=start_col + 3,
                                    min_row=header_row, 
                                    max_row=last_data_row)
            
            # Daily_Trucks (column E, with header)
            trucks_ref = Reference(self.ws, 
                                 min_col=start_col + 4,
                                 min_row=header_row, 
                                 max_row=last_data_row)
            
            # Get chart from chart module
            chart = self.receipt_charts.create_receipt_trend_chart(
                ws=self.ws,
                dates_ref=dates_ref,
                lines_ref=lines_ref,
                shipments_ref=shipments_ref,
                trucks_ref=trucks_ref
            )
            
            # Calculate chart position (to the right of table)
            chart_col = start_col + table_position['num_cols'] + columns_gap
            chart_position = f"{get_column_letter(chart_col)}{start_row}"
            
            # Add chart to worksheet
            self.ws.add_chart(chart, chart_position)
            
            print(f"✅ Receipt trend chart added at {chart_position}")
            return True
            
        except Exception as e:
            print(f"⚠️ Could not add receipt trend chart: {str(e)}")
            return False
    
    def add_receipt_volume_trend_chart(self, table_position: Dict, columns_gap: int = 2,
                                     row_gap: int = 22) -> bool:
        """
        Add Receipt Volume trend chart below the main receipt chart.
        
        Args:
            table_position: Dictionary with table location info
            columns_gap: Number of columns to leave between table and chart (default 2)
            row_gap: Number of rows below the first chart (default 22)
            
        Returns:
            bool: True if chart was added successfully, False otherwise
        """
        try:
            # Extract position info
            start_row = table_position['row']
            start_col = table_position['col']
            num_rows = table_position['num_rows']
            
            # Calculate rows
            header_row = start_row
            first_data_row = start_row + 1
            last_data_row = start_row + num_rows
            
            # Create References for the data
            # Dates (column A, no header)
            dates_ref = Reference(self.ws, 
                                 min_col=start_col, 
                                 min_row=first_data_row, 
                                 max_row=last_data_row)
            
            # Daily_Case_Equivalent_Volume (column H, with header)
            volume_ref = Reference(self.ws, 
                                  min_col=start_col + 7,  # Column H (8th column)
                                  min_row=header_row, 
                                  max_row=last_data_row)
            
            # Get chart from chart module
            chart = self.receipt_charts.create_receipt_volume_chart(
                ws=self.ws,
                dates_ref=dates_ref,
                volume_ref=volume_ref
            )
            
            # Calculate chart position (below the first chart)
            chart_col = start_col + table_position['num_cols'] + columns_gap
            chart_row = start_row + row_gap  # Position below first chart
            chart_position = f"{get_column_letter(chart_col)}{chart_row}"
            
            # Add chart to worksheet
            self.ws.add_chart(chart, chart_position)
            
            print(f"✅ Receipt volume chart added at {chart_position}")
            return True
            
        except Exception as e:
            print(f"⚠️ Could not add receipt volume chart: {str(e)}")
            return False
    
    def add_abc_fms_distribution_chart(self, table_position: Dict, columns_gap: int = 2) -> bool:
        """
        Add ABC-FMS distribution stacked bar chart to the right of the table.
        
        Args:
            table_position: Dictionary with table location info:
                - 'row': Starting row of table (1-based Excel indexing)  
                - 'col': Starting column of table (1-based)
                - 'num_rows': Number of data rows (excluding header)
                - 'num_cols': Number of columns in table
            columns_gap: Number of columns to leave between table and chart (default 2)
            
        Returns:
            bool: True if chart was added successfully, False otherwise
        """
        try:
            # Extract position info
            start_row = table_position['row']
            start_col = table_position['col']
            num_rows = table_position['num_rows']
            
            # Calculate chart data references
            # Categories (row headers: SKU, Volume, Lines)
            categories_ref = Reference(self.ws,
                                     min_col=start_col,
                                     min_row=start_row + 1,
                                     max_row=start_row + num_rows)
            
            # Data values (columns: AF, Rest, CS)
            data_ref = Reference(self.ws,
                               min_col=start_col + 1,
                               min_row=start_row,
                               max_col=start_col + table_position['num_cols'],
                               max_row=start_row + num_rows)
            
            # Get chart from chart module
            chart = self.abc_fms_charts.create_abc_fms_distribution_chart(
                ws=self.ws,
                data_ref=data_ref,
                categories_ref=categories_ref
            )
            
            # Calculate chart position (to the right of table)
            chart_col = start_col + table_position['num_cols'] + columns_gap
            chart_position = f"{get_column_letter(chart_col)}{start_row}"
            
            # Add chart to worksheet
            self.ws.add_chart(chart, chart_position)
            
            print(f"✅ ABC-FMS distribution chart added at {chart_position}")
            return True
            
        except Exception as e:
            print(f"⚠️ Could not add ABC-FMS distribution chart: {str(e)}")
            return False
    
    def add_sku_top_skus_chart(self, table_position: Dict, sku_col_idx: int,
                               volume_col_idx: int, num_skus: int = 10,
                               columns_gap: int = 2) -> bool:
        """
        Add horizontal bar chart of top N SKUs by volume to the SKU Analysis sheet.

        Args:
            table_position: Dictionary with table location info:
                - 'row': Header row of SKU table (1-based Excel indexing)
                - 'col': Starting column (1-based)
                - 'num_rows': Number of data rows (excluding header)
                - 'num_cols': Number of columns in table
            sku_col_idx: Column index (1-based) of SKU code column
            volume_col_idx: Column index (1-based) of volume data column
            num_skus: Number of top SKUs to chart (default 10)
            columns_gap: Gap between table and chart

        Returns:
            bool: Success status
        """
        try:
            start_row  = table_position['row']
            num_rows   = min(table_position['num_rows'], num_skus)
            header_row = start_row
            last_row   = start_row + num_rows

            # SKU codes (data rows only, no header — used as categories)
            sku_ref = Reference(self.ws, min_col=sku_col_idx,
                                min_row=header_row + 1, max_row=last_row)
            # Volume column — include header row for series name
            volume_ref = Reference(self.ws, min_col=volume_col_idx,
                                   min_row=header_row, max_row=last_row)

            chart = self.order_charts.create_top_skus_chart(
                self.ws, sku_ref, volume_ref, num_skus=num_skus
            )
            if chart:
                chart_position = self._calculate_chart_position(
                    table_position, placement='right', gap=columns_gap
                )
                self.ws.add_chart(chart, chart_position)

            print(f"✅ Top {num_skus} SKUs chart added")
            return True

        except Exception as e:
            print(f"⚠️ Could not add top SKUs chart: {str(e)}")
            return False

    # Helper methods
    def _create_reference(self, col: int, start_row: int, end_row: int, 
                         include_header: bool = False) -> Reference:
        """
        Helper method to create a Reference object.
        
        Args:
            col: Column number (1-based)
            start_row: Starting row (1-based)
            end_row: Ending row (1-based)
            include_header: Whether to include header row
            
        Returns:
            Reference object
        """
        if include_header:
            start_row -= 1
            
        return Reference(self.ws, 
                        min_col=col, 
                        min_row=start_row, 
                        max_row=end_row)
    
    def _calculate_chart_position(self, table_position: Dict, 
                                 placement: str = 'right',
                                 gap: int = 2) -> str:
        """
        Calculate where to place a chart relative to a table.
        
        Args:
            table_position: Table location dictionary
            placement: 'right', 'bottom', or 'below_right'
            gap: Number of rows/columns gap
            
        Returns:
            Cell reference string (e.g., 'J2')
        """
        if placement == 'right':
            col = table_position['col'] + table_position['num_cols'] + gap
            row = table_position['row']
        elif placement == 'bottom':
            col = table_position['col']
            row = table_position['row'] + table_position['num_rows'] + gap + 1
        elif placement == 'below_right':
            col = table_position['col'] + table_position['num_cols'] + gap
            row = table_position['row'] + table_position['num_rows'] + gap + 1
        else:
            # Default to right
            col = table_position['col'] + table_position['num_cols'] + gap
            row = table_position['row']
        return f"{get_column_letter(col)}{row}"

    # ── New charts: Inventory Analysis ───────────────────────────────────────

    def add_inventory_stock_trend_chart(self, daily_start_col: int, num_daily_rows: int,
                                        anchor: str) -> bool:
        """Line chart: daily total stock in cases over time.

        Args:
            daily_start_col: 1-based Excel column of daily_summary's Date column.
            num_daily_rows:  Number of data rows (excluding header).
            anchor:          Cell reference for chart top-left (e.g. 'AH1').
        """
        try:
            date_ref  = Reference(self.ws,
                                  min_col=daily_start_col,
                                  min_row=2, max_row=num_daily_rows + 1)
            cases_ref = Reference(self.ws,
                                  min_col=daily_start_col + 1,   # #Cases column
                                  min_row=1, max_row=num_daily_rows + 1)  # include header

            chart = LineChart()
            chart.title        = "Daily Total Stock (Cases)"
            chart.style        = 10
            chart.y_axis.title = "Cases"
            chart.x_axis.title = "Date"
            chart.width  = 20
            chart.height = 10
            chart.add_data(cases_ref, titles_from_data=True)
            chart.set_categories(date_ref)
            chart.series[0].graphicalProperties.line.solidFill = '4472C4'
            chart.series[0].graphicalProperties.line.width     = 20000

            self.ws.add_chart(chart, anchor)
            return True
        except Exception as e:
            print(f"Warning: Could not create inventory stock trend chart: {e}")
            return False

    def add_inventory_stock_status_chart(self, pie_start_col: int, pie_start_row: int,
                                         num_rows: int, anchor: str) -> bool:
        """Pie chart: SKU count by stock status (Low / Excess / OK / No Demand).

        Args:
            pie_start_col: 1-based Excel column of the 2-column helper table (Status | Count).
            pie_start_row: 1-based Excel row of the helper table header.
            num_rows:      Number of data rows (4 for the 4 statuses).
            anchor:        Cell reference for chart top-left.
        """
        try:
            labels_ref = Reference(self.ws,
                                   min_col=pie_start_col,
                                   min_row=pie_start_row + 1,
                                   max_row=pie_start_row + num_rows)
            data_ref   = Reference(self.ws,
                                   min_col=pie_start_col + 1,
                                   min_row=pie_start_row,       # include header for legend title
                                   max_row=pie_start_row + num_rows)

            chart = PieChart()
            chart.title  = "Stock Status Distribution (SKUs)"
            chart.style  = 10
            chart.width  = 12
            chart.height = 10
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(labels_ref)

            # Colour slices: Low=red, Excess=orange, OK=green, No Demand=gray
            colors = ['FF0000', 'ED7D31', '70AD47', 'A5A5A5']
            for i, color in enumerate(colors):
                pt = DataPoint(idx=i)
                pt.graphicalProperties.solidFill = color
                chart.series[0].dPt.append(pt)

            self.ws.add_chart(chart, anchor)
            return True
        except Exception as e:
            print(f"Warning: Could not create stock status pie chart: {e}")
            return False

    # ── New charts: Manpower Analysis ────────────────────────────────────────

    def add_manpower_pick_type_chart(self, table_position: Dict,
                                     columns_gap: int = 2) -> bool:
        """Column chart: order lines by pick type (Case / Each / Mixed / No Volume).

        Args:
            table_position: {'row', 'col', 'num_rows', 'num_cols'} — all 1-based Excel.
                            row = header row of the DataFrame.
            columns_gap:    Column gap between table and chart.
        """
        try:
            start_row = table_position['row']
            start_col = table_position['col']
            num_rows  = table_position['num_rows']

            cat_ref  = Reference(self.ws,
                                 min_col=start_col,
                                 min_row=start_row + 1,       # data rows only
                                 max_row=start_row + num_rows)
            data_ref = Reference(self.ws,
                                 min_col=start_col + 1,       # Order Lines column
                                 min_row=start_row,           # include header for legend
                                 max_row=start_row + num_rows)

            chart = BarChart()
            chart.type         = "col"
            chart.title        = "Pick Type Distribution — Order Lines"
            chart.y_axis.title = "Order Lines"
            chart.style = 10
            chart.width  = 15
            chart.height = 10
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cat_ref)
            chart.series[0].graphicalProperties.solidFill = '4472C4'

            anchor_col   = table_position['col'] + table_position['num_cols'] + columns_gap
            chart_anchor = f"{get_column_letter(anchor_col)}{start_row}"
            self.ws.add_chart(chart, chart_anchor)
            return True
        except Exception as e:
            print(f"Warning: Could not create pick type chart: {e}")
            return False

    def add_manpower_hourly_profile_chart(self, table_position: Dict,
                                          section_title: str = 'Picking',
                                          columns_gap: int = 2) -> bool:
        """Column chart: staff required by hour across the working day.

        Args:
            table_position: {'row', 'col', 'num_rows', 'num_cols'} — all 1-based Excel.
                            Columns expected: hour, ..., staff_required (last column).
            section_title:  'Picking' or 'Receiving' — used in chart title.
            columns_gap:    Column gap between table and chart.
        """
        try:
            start_row = table_position['row']
            start_col = table_position['col']
            num_rows  = table_position['num_rows']
            num_cols  = table_position['num_cols']

            cat_ref  = Reference(self.ws,
                                 min_col=start_col,              # hour column
                                 min_row=start_row + 1,
                                 max_row=start_row + num_rows)
            data_ref = Reference(self.ws,
                                 min_col=start_col + num_cols - 1,  # last col = staff_required
                                 min_row=start_row,              # include header for legend
                                 max_row=start_row + num_rows)

            chart = BarChart()
            chart.type         = "col"
            chart.title        = f"{section_title} — Hourly Staffing Profile"
            chart.y_axis.title = "Staff Required"
            chart.x_axis.title = "Hour"
            chart.style = 10
            chart.width  = 20
            chart.height = 10
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cat_ref)
            chart.series[0].graphicalProperties.solidFill = 'ED7D31'

            anchor_col   = start_col + num_cols + columns_gap
            chart_anchor = f"{get_column_letter(anchor_col)}{start_row}"
            self.ws.add_chart(chart, chart_anchor)
            return True
        except Exception as e:
            print(f"Warning: Could not create hourly staffing chart: {e}")
            return False

        return f"{get_column_letter(col)}{row}"